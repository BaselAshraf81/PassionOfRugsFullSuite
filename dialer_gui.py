#!/usr/bin/env python3
"""
Passion Of Rugs Advanced Dialer v4.1
Professional dialer with API integration and caching
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import pandas as pd
import requests
import base64
import threading
import webbrowser
import urllib.parse
from typing import Dict, List, Optional
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.datavalidation import DataValidation
from lead_processor_v2 import LeadProcessor
from cache_manager import CacheManager
from ai_assistant import AIAssistant, HAS_OPENAI
import os
import json
import datetime
import logging

logger = logging.getLogger(__name__)

# Try to import tkinterweb for HTML rendering
try:
    import tkinterweb
    HAS_TKINTERWEB = True
except ImportError:
    HAS_TKINTERWEB = False


class CloudTalkAPI:
    def __init__(self, access_key_id: str, access_key_secret: str):
        self.access_key_id = access_key_id
        self.access_key_secret = access_key_secret
        self.base_url = "https://my.cloudtalk.io/api"
        
        credentials = f"{access_key_id}:{access_key_secret}"
        encoded = base64.b64encode(credentials.encode()).decode()
        self.headers = {
            'Authorization': f'Basic {encoded}',
            'Content-Type': 'application/json'
        }
    
    def make_call(self, agent_id: str, callee_number: str) -> Dict:
        """Make a call via CloudTalk API"""
        url = f"{self.base_url}/calls/create.json"
        
        if not callee_number.startswith('+'):
            callee_number = '+' + callee_number.replace(' ', '').replace('(', '').replace(')', '').replace('-', '')
        
        payload = {"agent_id": agent_id, "callee_number": callee_number}
        
        try:
            response = requests.post(url, json=payload, headers=self.headers, timeout=10)
            
            if response.status_code == 200:
                return {"success": True, "message": "Call initiated successfully"}
            elif response.status_code == 403:
                return {"success": False, "message": "Agent is not online - cannot initiate call"}
            elif response.status_code == 406:
                return {"success": False, "message": "Invalid phone number format"}
            elif response.status_code == 409:
                return {"success": False, "message": "Agent already on a call"}
            else:
                return {"success": False, "message": f"Call failed - please try again"}
        except requests.exceptions.RequestException as e:
            return {"success": False, "message": f"Connection error: {str(e)}"}


class ExcelCache:
    """Manages caching of processed data in output Excel"""
    def __init__(self, output_file: str):
        self.output_file = output_file
        self.cache = {}
        self.load_cache()
    
    def load_cache(self):
        """Load existing data from output Excel"""
        if not os.path.exists(self.output_file):
            return
        
        try:
            df = pd.read_excel(self.output_file)
            for _, row in df.iterrows():
                orig_phone = str(row.get('Original Phone', '')).strip()
                if orig_phone:
                    # Split phones by comma, addresses by pipe
                    phones_raw = str(row.get('New Phone', ''))
                    addresses_raw = str(row.get('New Address', ''))
                    
                    # Check if this row was marked as address lookup failed (red background in Excel)
                    # We can infer this if there are phones but no addresses
                    has_phones = bool([p.strip() for p in phones_raw.split(',') if p.strip()] if pd.notna(row.get('New Phone')) else [])
                    has_addresses = bool([a.strip() for a in addresses_raw.split('|') if a.strip()] if pd.notna(row.get('New Address')) else [])
                    address_failed = has_phones and not has_addresses
                    
                    self.cache[orig_phone] = {
                        'original_address': row.get('Original Address', ''),
                        'original_phone': orig_phone,
                        'blank': '',
                        'age': row.get('Age', ''),
                        'original_name': row.get('Original Name', ''),
                        'new_phones': [p.strip() for p in phones_raw.split(',') if p.strip()] if pd.notna(row.get('New Phone')) else [],
                        'new_addresses': [a.strip() for a in addresses_raw.split('|') if a.strip()] if pd.notna(row.get('New Address')) else [],
                        'new_name': row.get('New Name', ''),
                        'status': row.get('Status', ''),
                        'notes': row.get('Notes', ''),
                        'address_lookup_failed': address_failed
                    }
        except Exception as e:
            print(f"Error loading cache: {e}")
    
    def get(self, original_phone: str) -> Optional[Dict]:
        return self.cache.get(original_phone)
    
    def update(self, original_phone: str, data: Dict):
        self.cache[original_phone] = data


class DialerGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Passion Of Rugs Advanced Dialer v4.1")
        self.root.geometry("1000x700")  # Optimized for standard screens
        self.root.minsize(900, 650)
        
        # Data
        self.input_excel_file = None
        self.output_excel_file = None
        self.excel_cache = None
        self.original_data = []
        self.current_person_idx = 0
        self.current_results = []
        self.current_result_idx = 0
        self.current_phone_idx = 0
        
        # APIs
        self.lead_processor = None
        self.cloudtalk_api = None
        self.agent_id = None
        self.ai_assistant = None
        
        # Settings (default values)
        self.settings = {
            'auto_phone_lookup': True,
            'auto_address_lookup': True,
            'enable_cloudtalk': True,
            'enable_cache': True,
            'enable_call_history': True,
            'ai_enabled': False,
            'ai_address_correction': True,
            'ai_person_filtering': True,
            'openai_api_key': ''
        }
        
        # AI data
        self.ai_results = None
        self.ai_correction_log = []
        
        # Call history
        self.call_history = []
        self.call_history_file = "call_history.json"
        self.load_call_history()
        
        # Call tracking for history
        self.last_called_phone = None
        self.last_called_name = None
        
        # Colors
        self.colors = {
            'bg': '#F5F7FA',
            'primary': '#2C3E50',
            'secondary': '#3498DB',
            'success': '#27AE60',
            'warning': '#F39C12',
            'danger': '#E74C3C',
            'light': '#ECF0F1',
            'white': '#FFFFFF'
        }
        
        self.root.configure(bg=self.colors['bg'])
        
        # Setup UI
        self.show_setup_screen()
    
    def load_call_history(self):
        """Load call history from JSON file"""
        if os.path.exists(self.call_history_file):
            try:
                with open(self.call_history_file, 'r') as f:
                    self.call_history = json.load(f)
            except:
                self.call_history = []
        else:
            self.call_history = []
    
    def save_call_history(self):
        """Save call history to JSON file"""
        try:
            with open(self.call_history_file, 'w') as f:
                json.dump(self.call_history, f, indent=2)
        except Exception as e:
            logger.error(f"Error saving call history: {e}")
    
    def add_to_call_history(self, phone, name, status, notes=""):
        """Add or update entry in call history - one entry per phone with most complete info"""
        if not self.settings.get('enable_call_history', True):
            return
        
        import datetime
        
        # Check if entry already exists for this phone
        existing_entry = None
        for entry in self.call_history:
            if entry.get('phone') == phone:
                existing_entry = entry
                break
        
        if existing_entry:
            # Update existing entry with more complete information
            existing_entry['timestamp'] = datetime.datetime.now().isoformat()
            if name:
                existing_entry['name'] = name
            if status:
                existing_entry['status'] = status
            if notes:
                existing_entry['notes'] = notes
        else:
            # Create new entry
            entry = {
                'timestamp': datetime.datetime.now().isoformat(),
                'phone': phone,
                'name': name,
                'status': status,
                'notes': notes
            }
            self.call_history.append(entry)
        
        self.save_call_history()
    
    def show_setup_screen(self):
        """Show setup configuration screen with scrollable content"""
        # Main container
        main_container = tk.Frame(self.root, bg=self.colors['bg'])
        main_container.pack(fill=tk.BOTH, expand=True)
        
        # Title (fixed at top)
        title = tk.Label(
            main_container,
            text="Passion Of Rugs Advanced Dialer v4.1 - Setup",
            font=('Arial', 16, 'bold'),
            bg=self.colors['bg'],
            fg=self.colors['primary']
        )
        title.pack(pady=(10, 10))
        
        # Create canvas with scrollbar for scrollable content
        canvas_frame = tk.Frame(main_container, bg=self.colors['bg'])
        canvas_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=(0, 10))
        
        canvas = tk.Canvas(canvas_frame, bg=self.colors['bg'], highlightthickness=0)
        scrollbar = ttk.Scrollbar(canvas_frame, orient="vertical", command=canvas.yview)
        
        setup_frame = tk.Frame(canvas, bg=self.colors['bg'])
        
        canvas.create_window((0, 0), window=setup_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Enable mouse wheel scrolling
        def on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        canvas.bind_all("<MouseWheel>", on_mousewheel)
        
        # Form frame
        form = tk.Frame(setup_frame, bg=self.colors['white'], relief=tk.RAISED, bd=2)
        form.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        # Grid configuration
        form.grid_columnconfigure(1, weight=1)
        
        row = 0
        
        # TrestleIQ API Key
        tk.Label(form, text="TrestleIQ API Key:", font=('Arial', 9, 'bold'), bg=self.colors['white']).grid(
            row=row, column=0, sticky='e', padx=10, pady=8
        )
        self.api_key_entry = tk.Entry(form, font=('Arial', 9), width=45)
        self.api_key_entry.grid(row=row, column=1, sticky='ew', padx=10, pady=8)
        row += 1
        
        # Separator
        ttk.Separator(form, orient='horizontal').grid(row=row, column=0, columnspan=2, sticky='ew', pady=5)
        row += 1
        
        # CloudTalk section
        tk.Label(form, text="CloudTalk Configuration", font=('Arial', 10, 'bold'), bg=self.colors['white'], 
                fg=self.colors['secondary']).grid(row=row, column=0, columnspan=2, pady=5)
        row += 1
        
        tk.Label(form, text="Access Key ID:", font=('Arial', 9), bg=self.colors['white']).grid(
            row=row, column=0, sticky='e', padx=10, pady=5
        )
        self.cloudtalk_id_entry = tk.Entry(form, font=('Arial', 9), width=45)
        self.cloudtalk_id_entry.grid(row=row, column=1, sticky='ew', padx=10, pady=5)
        row += 1
        
        tk.Label(form, text="Access Key Secret:", font=('Arial', 9), bg=self.colors['white']).grid(
            row=row, column=0, sticky='e', padx=10, pady=5
        )
        self.cloudtalk_secret_entry = tk.Entry(form, font=('Arial', 9), width=45, show='*')
        self.cloudtalk_secret_entry.grid(row=row, column=1, sticky='ew', padx=10, pady=5)
        row += 1
        
        tk.Label(form, text="Agent ID:", font=('Arial', 9), bg=self.colors['white']).grid(
            row=row, column=0, sticky='e', padx=10, pady=5
        )
        self.agent_id_entry = tk.Entry(form, font=('Arial', 9), width=45)
        self.agent_id_entry.grid(row=row, column=1, sticky='ew', padx=10, pady=5)
        row += 1
        
        # Default config button
        default_btn = tk.Button(
            form,
            text="Use Default Config",
            command=self.use_default_cloudtalk_config,
            font=('Arial', 8),
            bg=self.colors['secondary'],
            fg='white',
            padx=10,
            pady=4,
            cursor='hand2'
        )
        default_btn.grid(row=row, column=1, sticky='w', padx=10, pady=5)
        row += 1
        
        # Separator
        ttk.Separator(form, orient='horizontal').grid(row=row, column=0, columnspan=2, sticky='ew', pady=5)
        row += 1
        
        # Input file
        tk.Label(form, text="Input Excel File:", font=('Arial', 9, 'bold'), bg=self.colors['white']).grid(
            row=row, column=0, sticky='e', padx=10, pady=8
        )
        file_frame = tk.Frame(form, bg=self.colors['white'])
        file_frame.grid(row=row, column=1, sticky='ew', padx=10, pady=8)
        
        self.input_file_label = tk.Label(file_frame, text="No file selected", font=('Arial', 9), 
                                         bg=self.colors['white'], fg='gray')
        self.input_file_label.pack(side=tk.LEFT, padx=(0, 5))
        
        browse_btn = tk.Button(
            file_frame,
            text="Browse...",
            command=self.browse_input_file,
            font=('Arial', 8),
            bg=self.colors['light'],
            padx=10,
            pady=3,
            cursor='hand2'
        )
        browse_btn.pack(side=tk.LEFT)
        row += 1
        
        # Output file
        tk.Label(form, text="Output File Name:", font=('Arial', 9, 'bold'), bg=self.colors['white']).grid(
            row=row, column=0, sticky='e', padx=10, pady=8
        )
        self.output_file_entry = tk.Entry(form, font=('Arial', 9), width=45)
        self.output_file_entry.insert(0, "processed_leads.xlsx")
        self.output_file_entry.grid(row=row, column=1, sticky='ew', padx=10, pady=8)
        row += 1
        
        # Separator
        ttk.Separator(form, orient='horizontal').grid(row=row, column=0, columnspan=2, sticky='ew', pady=5)
        row += 1
        
        # AI Configuration section
        tk.Label(form, text="AI Configuration (Optional)", font=('Arial', 10, 'bold'), bg=self.colors['white'], 
                fg=self.colors['secondary']).grid(row=row, column=0, columnspan=2, pady=5)
        row += 1
        
        tk.Label(form, text="OpenAI API Key:", font=('Arial', 9), bg=self.colors['white']).grid(
            row=row, column=0, sticky='e', padx=10, pady=5
        )
        ai_key_frame = tk.Frame(form, bg=self.colors['white'])
        ai_key_frame.grid(row=row, column=1, sticky='ew', padx=10, pady=5)
        
        self.openai_key_entry = tk.Entry(ai_key_frame, font=('Arial', 9), width=25, show='*')
        self.openai_key_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        
        self.use_default_ai_btn = tk.Button(
            ai_key_frame,
            text="Use Default Config",
            command=self.use_default_openai_config,
            font=('Arial', 8),
            bg=self.colors['secondary'],
            fg='white',
            padx=8,
            pady=2,
            cursor='hand2'
        )
        self.use_default_ai_btn.pack(side=tk.LEFT, padx=(0, 5))
        
        self.test_ai_btn = tk.Button(
            ai_key_frame,
            text="Test",
            command=self.test_ai_connection,
            font=('Arial', 8),
            bg=self.colors['secondary'],
            fg='white',
            padx=8,
            pady=2,
            cursor='hand2'
        )
        self.test_ai_btn.pack(side=tk.LEFT)
        row += 1
        
        # AI info
        ai_info = tk.Label(
            form,
            text="AI features: Address correction & intelligent person filtering (GPT-5 nano - $0.025/$0.20 per 1M tokens)",
            font=('Arial', 8, 'italic'),
            bg=self.colors['white'],
            fg='#666',
            wraplength=450,
            justify='left'
        )
        ai_info.grid(row=row, column=0, columnspan=2, padx=10, pady=(0, 8))
        row += 1
        
        # Separator
        ttk.Separator(form, orient='horizontal').grid(row=row, column=0, columnspan=2, sticky='ew', pady=5)
        row += 1
        
        # Cache management section
        tk.Label(form, text="Cache Management", font=('Arial', 10, 'bold'), bg=self.colors['white'], 
                fg=self.colors['secondary']).grid(row=row, column=0, columnspan=2, pady=5)
        row += 1
        
        cache_btn_frame = tk.Frame(form, bg=self.colors['white'])
        cache_btn_frame.grid(row=row, column=0, columnspan=2, pady=8)
        
        tk.Button(
            cache_btn_frame,
            text="📊 Cache Stats",
            command=self.show_cache_stats,
            font=('Arial', 8),
            bg=self.colors['secondary'],
            fg='white',
            padx=10,
            pady=5,
            cursor='hand2'
        ).pack(side=tk.LEFT, padx=5)
        
        tk.Button(
            cache_btn_frame,
            text="🗑️ Clear Cache",
            command=self.clear_cache_confirm,
            font=('Arial', 8),
            bg=self.colors['warning'],
            fg='white',
            padx=10,
            pady=5,
            cursor='hand2'
        ).pack(side=tk.LEFT, padx=5)
        row += 1
        
        # Settings and Start buttons
        button_frame = tk.Frame(form, bg=self.colors['white'])
        button_frame.grid(row=row, column=0, columnspan=2, pady=15)
        
        settings_btn = tk.Button(
            button_frame,
            text="⚙️ Settings",
            command=self.show_settings_window,
            font=('Arial', 10, 'bold'),
            bg=self.colors['secondary'],
            fg='white',
            padx=20,
            pady=8,
            cursor='hand2'
        )
        settings_btn.pack(side=tk.LEFT, padx=5)
        
        start_btn = tk.Button(
            button_frame,
            text="Start Dialer",
            command=lambda: self.start_dialer(main_container),
            font=('Arial', 11, 'bold'),
            bg=self.colors['success'],
            fg='white',
            padx=30,
            pady=10,
            cursor='hand2'
        )
        start_btn.pack(side=tk.LEFT, padx=5)
        
        # Update scroll region
        setup_frame.update_idletasks()
        canvas.config(scrollregion=canvas.bbox("all"))
        
        # Bind mousewheel to canvas (local binding, not global)
        def _on_mousewheel(event):
            try:
                canvas.yview_scroll(int(-1*(event.delta/120)), "units")
            except tk.TclError:
                # Canvas has been destroyed, ignore the event
                pass
        canvas.bind("<MouseWheel>", _on_mousewheel)
        
        # Also bind to the canvas frame to catch mouse events
        canvas_frame.bind("<MouseWheel>", _on_mousewheel)
        
        # Try to load from config
        try:
            from config import API_KEY
            self.api_key_entry.insert(0, API_KEY)
        except:
            pass
        
        # Try to load OpenAI API key from config
        try:
            from config import OPENAI_API_KEY
            if OPENAI_API_KEY:
                self.openai_key_entry.insert(0, OPENAI_API_KEY)
        except:
            pass
    
    def use_default_cloudtalk_config(self):
        """Fill in default CloudTalk configuration"""
        try:
            from config import CLOUDTALK_ACCESS_KEY_ID, CLOUDTALK_ACCESS_KEY_SECRET, CLOUDTALK_AGENT_ID
            if CLOUDTALK_ACCESS_KEY_ID:
                self.cloudtalk_id_entry.delete(0, tk.END)
                self.cloudtalk_id_entry.insert(0, CLOUDTALK_ACCESS_KEY_ID)
            if CLOUDTALK_ACCESS_KEY_SECRET:
                self.cloudtalk_secret_entry.delete(0, tk.END)
                self.cloudtalk_secret_entry.insert(0, CLOUDTALK_ACCESS_KEY_SECRET)
            if CLOUDTALK_AGENT_ID:
                self.agent_id_entry.delete(0, tk.END)
                self.agent_id_entry.insert(0, CLOUDTALK_AGENT_ID)
            if CLOUDTALK_ACCESS_KEY_ID or CLOUDTALK_ACCESS_KEY_SECRET or CLOUDTALK_AGENT_ID:
                messagebox.showinfo("Default Config", "Default CloudTalk configuration loaded from config.py")
            else:
                messagebox.showwarning("No Default Config", 
                                     "No default CloudTalk configuration found in config.py.\n"
                                     "Please add your credentials to config.py")
        except ImportError:
            messagebox.showerror("Config Error", 
                               "config.py not found. Please create it with your CloudTalk credentials.")
    
    def use_default_openai_config(self):
        """Fill in default OpenAI API key configuration"""
        try:
            from config import OPENAI_API_KEY
            if OPENAI_API_KEY:
                self.openai_key_entry.delete(0, tk.END)
                self.openai_key_entry.insert(0, OPENAI_API_KEY)
                messagebox.showinfo("Default Config", "Default OpenAI API key loaded from config.py")
            else:
                messagebox.showwarning("No Default Config", 
                                     "No default OpenAI API key found in config.py.\n"
                                     "Please add your API key to OPENAI_API_KEY in config.py")
        except ImportError:
            messagebox.showerror("Config Error", 
                               "config.py not found. Please create it with your OpenAI API key.")
    
    def test_ai_connection(self):
        """Test OpenAI API connection"""
        api_key = self.openai_key_entry.get().strip()
        
        if not api_key:
            messagebox.showwarning("No API Key", "Please enter an OpenAI API key first")
            return
        
        if not HAS_OPENAI:
            messagebox.showerror("OpenAI Not Installed", 
                               "OpenAI library not installed. Run: pip install openai")
            return
        
        # Test connection
        self.test_ai_btn.config(state='disabled', text='Testing...')
        self.root.update()
        
        try:
            test_ai = AIAssistant(api_key)
            success, message = test_ai.test_connection()
            
            if success:
                messagebox.showinfo("Connection Successful", 
                                  "OpenAI API connection successful!\nAI features will be enabled.")
            else:
                messagebox.showerror("Connection Failed", message)
        finally:
            self.test_ai_btn.config(state='normal', text='Test')
    
    def show_settings_window(self):
        """Show settings configuration window"""
        settings_win = tk.Toplevel(self.root)
        settings_win.title("Settings")
        # Increase height if AI is available
        height = "350" if self.ai_assistant else "250"
        settings_win.geometry(f"400x{height}")
        settings_win.configure(bg=self.colors['bg'])
        settings_win.transient(self.root)
        settings_win.grab_set()
        
        # Title
        tk.Label(
            settings_win,
            text="API Settings",
            font=('Arial', 13, 'bold'),
            bg=self.colors['bg'],
            fg=self.colors['primary']
        ).pack(pady=15)
        
        # Settings frame
        settings_frame = tk.Frame(settings_win, bg=self.colors['white'], relief=tk.RAISED, bd=2)
        settings_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=5)
        
        self.setting_vars = {}
        
        # Info text
        tk.Label(
            settings_frame,
            text="Control automatic API lookups when loading a person:",
            font=('Arial', 9),
            bg=self.colors['white'],
            fg='#666',
            wraplength=320,
            justify='left'
        ).pack(pady=(15, 10), padx=15)
        
        # Auto Phone Lookup
        var = tk.BooleanVar(value=self.settings.get('auto_phone_lookup', True))
        self.setting_vars['auto_phone_lookup'] = var
        tk.Checkbutton(
            settings_frame,
            text="Auto Phone Lookup",
            variable=var,
            font=('Arial', 10, 'bold'),
            bg=self.colors['white'],
            activebackground=self.colors['white']
        ).pack(anchor='w', padx=25, pady=8)
        
        # Auto Address Lookup
        var = tk.BooleanVar(value=self.settings.get('auto_address_lookup', True))
        self.setting_vars['auto_address_lookup'] = var
        tk.Checkbutton(
            settings_frame,
            text="Auto Address Lookup",
            variable=var,
            font=('Arial', 10, 'bold'),
            bg=self.colors['white'],
            activebackground=self.colors['white']
        ).pack(anchor='w', padx=25, pady=8)
        
        # AI Settings section (if AI is available)
        if self.ai_assistant:
            ttk.Separator(settings_frame, orient='horizontal').pack(fill=tk.X, pady=10)
            
            tk.Label(
                settings_frame,
                text="AI Features:",
                font=('Arial', 9),
                bg=self.colors['white'],
                fg='#666'
            ).pack(pady=(5, 5), padx=15, anchor='w')
            
            # AI Address Correction
            var = tk.BooleanVar(value=self.settings.get('ai_address_correction', True))
            self.setting_vars['ai_address_correction'] = var
            tk.Checkbutton(
                settings_frame,
                text="AI Address Correction (auto-retry on failure)",
                variable=var,
                font=('Arial', 9),
                bg=self.colors['white'],
                activebackground=self.colors['white']
            ).pack(anchor='w', padx=25, pady=5)
            
            # AI Person Filtering
            var = tk.BooleanVar(value=self.settings.get('ai_person_filtering', True))
            self.setting_vars['ai_person_filtering'] = var
            tk.Checkbutton(
                settings_frame,
                text="AI Person Filtering & Insights",
                variable=var,
                font=('Arial', 9),
                bg=self.colors['white'],
                activebackground=self.colors['white']
            ).pack(anchor='w', padx=25, pady=5)
        
        # Note
        tk.Label(
            settings_frame,
            text="Note: Manual buttons always work regardless of these settings",
            font=('Arial', 8, 'italic'),
            bg=self.colors['white'],
            fg='#999'
        ).pack(pady=(10, 15), padx=15)
        
        # Buttons
        button_frame = tk.Frame(settings_win, bg=self.colors['bg'])
        button_frame.pack(pady=12)
        
        tk.Button(
            button_frame,
            text="Save Settings",
            command=lambda: self.save_settings(settings_win),
            font=('Arial', 9, 'bold'),
            bg=self.colors['success'],
            fg='white',
            padx=20,
            pady=6,
            cursor='hand2'
        ).pack(side=tk.LEFT, padx=5)
        
        tk.Button(
            button_frame,
            text="Cancel",
            command=settings_win.destroy,
            font=('Arial', 9),
            bg=self.colors['light'],
            padx=20,
            pady=6,
            cursor='hand2'
        ).pack(side=tk.LEFT, padx=5)
    
    def save_settings(self, window):
        """Save settings from settings window"""
        for key, var in self.setting_vars.items():
            self.settings[key] = var.get()
        
        # Save to file
        try:
            with open('dialer_settings.json', 'w') as f:
                json.dump(self.settings, f, indent=2)
        except:
            pass
        
        # Only update status if status_bar exists (in dialer screen)
        if hasattr(self, 'status_bar'):
            self.update_status("Settings saved successfully", self.colors['success'])
        
        messagebox.showinfo("Settings Saved", "Settings have been saved successfully.")
        window.destroy()
    
    def load_settings(self):
        """Load settings from file"""
        try:
            if os.path.exists('dialer_settings.json'):
                with open('dialer_settings.json', 'r') as f:
                    loaded = json.load(f)
                    self.settings.update(loaded)
        except:
            pass
    
    def browse_input_file(self):
        """Browse for input Excel file"""
        filename = filedialog.askopenfilename(
            title="Select Input Excel File",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if filename:
            self.input_excel_file = filename
            self.input_file_label.config(text=os.path.basename(filename), fg='black')
    
    def start_dialer(self, setup_frame):
        """Initialize and start the dialer"""
        # Load settings
        self.load_settings()
        
        api_key = self.api_key_entry.get().strip()
        cloudtalk_id = self.cloudtalk_id_entry.get().strip()
        cloudtalk_secret = self.cloudtalk_secret_entry.get().strip()
        agent_id = self.agent_id_entry.get().strip()
        
        if not api_key:
            messagebox.showerror("Error", "Please enter TrestleIQ API Key")
            return
        
        if not self.input_excel_file:
            messagebox.showerror("Error", "Please select an input Excel file")
            return
        
        output_filename = self.output_file_entry.get().strip()
        if not output_filename:
            messagebox.showerror("Error", "Please enter an output Excel file name")
            return
        
        # Create output file path
        input_dir = os.path.dirname(self.input_excel_file)
        self.output_excel_file = os.path.join(input_dir, output_filename)
        if not self.output_excel_file.endswith('.xlsx'):
            self.output_excel_file += '.xlsx'
        
        # Check if output file exists and ask user
        if os.path.exists(self.output_excel_file):
            response = messagebox.askyesnocancel(
                "Existing File Found",
                f"The output file '{output_filename}' already exists.\n\n"
                "Yes: Continue with existing file (load cached data)\n"
                "No: Create new file (overwrite existing)\n"
                "Cancel: Go back"
            )
            
            if response is None:  # Cancel
                return
            elif response is False:  # No - create new
                try:
                    os.remove(self.output_excel_file)
                except:
                    pass
        
        # Initialize AI Assistant first
        openai_key = self.openai_key_entry.get().strip()
        if openai_key and HAS_OPENAI:
            self.ai_assistant = AIAssistant(openai_key)
            self.settings['ai_enabled'] = True
            self.settings['openai_api_key'] = openai_key
        else:
            self.ai_assistant = None
            self.settings['ai_enabled'] = False
        
        # Initialize APIs with AI assistant
        use_cache = self.settings.get('enable_cache', True)
        self.lead_processor = LeadProcessor(api_key, use_cache=use_cache, ai_assistant=self.ai_assistant)
        
        if cloudtalk_id and cloudtalk_secret and agent_id:
            if self.settings.get('enable_cloudtalk', True):
                self.cloudtalk_api = CloudTalkAPI(cloudtalk_id, cloudtalk_secret)
                self.agent_id = agent_id
        
        # Initialize cache
        self.excel_cache = ExcelCache(self.output_excel_file)
        
        # Load data
        try:
            df = pd.read_excel(self.input_excel_file)
            self.original_data = df.to_dict('records')
            
            # Hide setup, show dialer
            setup_frame.destroy()
            self.show_dialer_screen()
            
            # Load first person
            self.load_person(0)
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load Excel file: {str(e)}")
    
    def show_dialer_screen(self):
        """Show main dialer interface with proper grid layout and integrated settings"""
        # Main container
        main = tk.Frame(self.root, bg=self.colors['bg'])
        main.pack(fill=tk.BOTH, expand=True)
        
        # Configure grid weights - add column for settings
        main.grid_rowconfigure(1, weight=0)  # Status bar
        main.grid_rowconfigure(2, weight=0)  # Original person
        main.grid_rowconfigure(3, weight=1)  # Results (expandable)
        main.grid_rowconfigure(4, weight=0)  # Navigation
        main.grid_columnconfigure(0, weight=3)  # Main content (75%)
        main.grid_columnconfigure(1, weight=1)  # Settings panel (25%)
        
        # Title with settings toggle
        title_frame = tk.Frame(main, bg=self.colors['primary'], height=30)
        title_frame.grid(row=0, column=0, columnspan=2, sticky='ew')
        title_frame.grid_propagate(False)
        
        title_content = tk.Frame(title_frame, bg=self.colors['primary'])
        title_content.pack(expand=True, fill=tk.X)
        
        tk.Label(
            title_content,
            text="Passion Of Rugs Advanced Dialer v4.1",
            font=('Arial', 11, 'bold'),
            bg=self.colors['primary'],
            fg='white'
        ).pack(side=tk.LEFT, padx=10)
        
        # Settings toggle button
        self.settings_visible = tk.BooleanVar(value=True)
        tk.Checkbutton(
            title_content,
            text="⚙️ Settings",
            variable=self.settings_visible,
            command=self.toggle_settings_panel,
            font=('Arial', 9),
            bg=self.colors['primary'],
            fg='white',
            selectcolor=self.colors['primary'],
            activebackground=self.colors['primary'],
            activeforeground='white',
            cursor='hand2'
        ).pack(side=tk.RIGHT, padx=10)
        
        # Comprehensive Status bar
        status_frame = tk.Frame(main, bg=self.colors['light'], relief=tk.SUNKEN, bd=1)
        status_frame.grid(row=1, column=0, columnspan=2, sticky='ew', padx=5, pady=(3, 2))
        
        # Status sections
        self.status_api = tk.Label(
            status_frame,
            text="API: Ready",
            font=('Arial', 8),
            bg=self.colors['light'],
            fg=self.colors['primary'],
            anchor='w',
            padx=5
        )
        self.status_api.pack(side=tk.LEFT)
        
        ttk.Separator(status_frame, orient='vertical').pack(side=tk.LEFT, fill=tk.Y, padx=5)
        
        self.status_cache = tk.Label(
            status_frame,
            text="Cache: Enabled",
            font=('Arial', 8),
            bg=self.colors['light'],
            fg=self.colors['primary'],
            anchor='w',
            padx=5
        )
        self.status_cache.pack(side=tk.LEFT)
        
        ttk.Separator(status_frame, orient='vertical').pack(side=tk.LEFT, fill=tk.Y, padx=5)
        
        self.status_bar = tk.Label(
            status_frame,
            text="Ready",
            font=('Arial', 8),
            bg=self.colors['light'],
            fg=self.colors['primary'],
            anchor='w',
            padx=5
        )
        self.status_bar.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        # Original Person Section
        orig_frame = tk.LabelFrame(
            main,
            text="ORIGINAL PERSON",
            font=('Arial', 9, 'bold'),
            bg=self.colors['white'],
            fg=self.colors['primary'],
            relief=tk.RAISED,
            bd=1
        )
        orig_frame.grid(row=2, column=0, sticky='ew', padx=5, pady=2)
        
        orig_grid = tk.Frame(orig_frame, bg=self.colors['white'])
        orig_grid.pack(fill=tk.BOTH, padx=8, pady=5)
        
        self.orig_labels = {}
        
        # Original Name
        tk.Label(orig_grid, text="Name:", font=('Arial', 8, 'bold'), 
                bg=self.colors['white']).grid(row=0, column=0, sticky='e', padx=(0, 3), pady=2)
        entry = tk.Entry(orig_grid, font=('Arial', 8), relief=tk.FLAT, 
                        bg=self.colors['white'], fg=self.colors['primary'], 
                        readonlybackground=self.colors['white'], state='readonly')
        entry.grid(row=0, column=1, sticky='ew', padx=(0, 10), pady=2)
        self.orig_labels['Original Name'] = entry
        
        # Original Phone
        tk.Label(orig_grid, text="Phone:", font=('Arial', 8, 'bold'), 
                bg=self.colors['white']).grid(row=0, column=2, sticky='e', padx=(0, 3), pady=2)
        entry = tk.Entry(orig_grid, font=('Arial', 8), relief=tk.FLAT, 
                        bg=self.colors['white'], fg=self.colors['primary'],
                        readonlybackground=self.colors['white'], state='readonly', width=15)
        entry.grid(row=0, column=3, sticky='ew', padx=(0, 10), pady=2)
        self.orig_labels['Original Phone'] = entry
        
        # Age
        tk.Label(orig_grid, text="Age:", font=('Arial', 8, 'bold'), 
                bg=self.colors['white']).grid(row=0, column=4, sticky='e', padx=(0, 3), pady=2)
        entry = tk.Entry(orig_grid, font=('Arial', 8), relief=tk.FLAT, 
                        bg=self.colors['white'], fg=self.colors['primary'],
                        readonlybackground=self.colors['white'], state='readonly', width=5)
        entry.grid(row=0, column=5, sticky='w', pady=2)
        self.orig_labels['Age'] = entry
        
        # Original Address with buttons
        tk.Label(orig_grid, text="Address:", font=('Arial', 8, 'bold'), 
                bg=self.colors['white']).grid(row=1, column=0, sticky='e', padx=(0, 3), pady=2)
        
        orig_addr_container = tk.Frame(orig_grid, bg=self.colors['white'])
        orig_addr_container.grid(row=1, column=1, columnspan=5, sticky='ew', pady=2)
        
        entry = tk.Entry(orig_addr_container, font=('Arial', 8), relief=tk.FLAT, 
                        bg=self.colors['white'], fg=self.colors['primary'],
                        readonlybackground=self.colors['white'], state='readonly')
        entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 3))
        self.orig_labels['Original Address'] = entry
        
        tk.Button(
            orig_addr_container,
            text="📋",
            command=self.copy_original_address,
            font=('Arial', 7),
            bg=self.colors['light'],
            padx=3,
            pady=1,
            cursor='hand2'
        ).pack(side=tk.LEFT, padx=1)
        
        tk.Button(
            orig_addr_container,
            text="🔍",
            command=self.google_search_original_address,
            font=('Arial', 7),
            bg=self.colors['secondary'],
            fg='white',
            padx=3,
            pady=1,
            cursor='hand2'
        ).pack(side=tk.LEFT, padx=1)
        
        orig_grid.grid_columnconfigure(1, weight=1)
        
        # Results Section with Tabs
        result_frame = tk.LabelFrame(
            main,
            text="RESULTS",
            font=('Arial', 9, 'bold'),
            bg=self.colors['white'],
            fg=self.colors['primary'],
            relief=tk.RAISED,
            bd=1
        )
        result_frame.grid(row=3, column=0, sticky='nsew', padx=5, pady=2)
        
        # Create notebook for tabs
        self.results_notebook = ttk.Notebook(result_frame)
        self.results_notebook.pack(fill=tk.BOTH, expand=True, padx=2, pady=2)
        
        # Tab 1: AI Overview (Default tab)
        self.ai_tab = tk.Frame(self.results_notebook, bg='white')
        self.results_notebook.add(self.ai_tab, text="AI Overview & Filtering")
        
        # Tab 2: Standard Results
        standard_tab = tk.Frame(self.results_notebook, bg=self.colors['white'])
        self.results_notebook.add(standard_tab, text="Standard View")
        
        # Now create the standard results view in standard_tab instead of result_frame
        result_frame_inner = standard_tab
        
        # Result header with counter and last name match
        result_header = tk.Frame(result_frame_inner, bg=self.colors['secondary'], height=25)
        result_header.pack(fill=tk.X)
        result_header.pack_propagate(False)
        
        self.result_counter_label = tk.Label(
            result_header,
            text="Result 0 of 0",
            font=('Arial', 8, 'bold'),
            bg=self.colors['secondary'],
            fg='white'
        )
        self.result_counter_label.pack(side=tk.LEFT, padx=10)
        
        self.lastname_match_label = tk.Label(
            result_header,
            text="",
            font=('Arial', 9, 'bold'),
            bg=self.colors['warning'],
            fg='white',
            padx=10,
            pady=3
        )
        
        self.address_warning_label = tk.Label(
            result_header,
            text="⚠️ ADDRESS LOOKUP FAILED",
            font=('Arial', 9, 'bold'),
            bg=self.colors['danger'],
            fg='white',
            padx=10,
            pady=3
        )
        
        # Result content (no scrolling)
        self.result_content = tk.Frame(result_frame_inner, bg=self.colors['white'])
        self.result_content.pack(fill=tk.BOTH, expand=True, padx=10, pady=8)
        
        # Result content grid
        self.result_content.grid_columnconfigure(1, weight=1)
        
        # New Name and Age on same row
        name_age_frame = tk.Frame(self.result_content, bg=self.colors['white'])
        name_age_frame.grid(row=0, column=0, columnspan=2, sticky='ew', padx=5, pady=2)
        
        tk.Label(name_age_frame, text="Name:", font=('Arial', 8, 'bold'), 
                bg=self.colors['white']).pack(side=tk.LEFT, padx=(0, 3))
        self.new_name_entry = tk.Entry(name_age_frame, font=('Arial', 8), relief=tk.FLAT,
                                       bg=self.colors['white'], fg=self.colors['primary'],
                                       readonlybackground=self.colors['white'], state='readonly')
        self.new_name_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        
        tk.Label(name_age_frame, text="Age:", font=('Arial', 8, 'bold'), 
                bg=self.colors['white']).pack(side=tk.LEFT, padx=(0, 3))
        self.age_entry = tk.Entry(name_age_frame, font=('Arial', 8), relief=tk.FLAT,
                                 bg=self.colors['white'], fg=self.colors['primary'],
                                 readonlybackground=self.colors['white'], state='readonly', width=5)
        self.age_entry.pack(side=tk.LEFT)
        
        # Phone with buttons
        phone_frame = tk.Frame(self.result_content, bg=self.colors['white'])
        phone_frame.grid(row=1, column=0, columnspan=2, sticky='ew', padx=5, pady=2)
        
        tk.Label(phone_frame, text="Phone:", font=('Arial', 8, 'bold'), 
                bg=self.colors['white']).pack(side=tk.LEFT, padx=(0, 3))
        
        self.phone_entry = tk.Entry(phone_frame, font=('Arial', 8, 'bold'), relief=tk.FLAT,
                                    bg=self.colors['white'], fg=self.colors['primary'],
                                    readonlybackground=self.colors['white'], state='readonly', width=15)
        self.phone_entry.pack(side=tk.LEFT, padx=(0, 5))
        
        self.copy_btn = tk.Button(
            phone_frame,
            text="📋",
            command=self.copy_phone,
            font=('Arial', 7),
            bg=self.colors['light'],
            padx=3,
            pady=1,
            cursor='hand2'
        )
        self.copy_btn.pack(side=tk.LEFT, padx=1)
        
        self.call_btn = tk.Button(
            phone_frame,
            text="📞 Call",
            command=self.make_call,
            font=('Arial', 8, 'bold'),
            bg=self.colors['success'],
            fg='white',
            padx=6,
            pady=2,
            cursor='hand2'
        )
        self.call_btn.pack(side=tk.LEFT, padx=2)
        
        # Refresh button
        self.refresh_btn = tk.Button(
            phone_frame,
            text="🔄",
            command=self.refresh_current_person,
            font=('Arial', 8),
            bg=self.colors['warning'],
            fg='white',
            padx=4,
            pady=2,
            cursor='hand2'
        )
        self.refresh_btn.pack(side=tk.LEFT, padx=2)
        
        # Full Response button
        self.full_response_btn = tk.Button(
            phone_frame,
            text="📄",
            command=self.show_full_response,
            font=('Arial', 8),
            bg=self.colors['secondary'],
            fg='white',
            padx=4,
            pady=2,
            cursor='hand2'
        )
        self.full_response_btn.pack(side=tk.LEFT, padx=2)
        
        # Phone navigation
        self.phone_nav_frame = tk.Frame(phone_frame, bg=self.colors['white'])
        self.phone_nav_frame.pack(side=tk.LEFT, padx=10)
        
        self.prev_phone_btn = tk.Button(
            self.phone_nav_frame,
            text="◄",
            command=self.prev_phone,
            font=('Arial', 8),
            bg=self.colors['secondary'],
            fg='white',
            padx=5,
            pady=1,
            cursor='hand2'
        )
        
        self.phone_counter_label = tk.Label(
            self.phone_nav_frame,
            text="",
            font=('Arial', 8),
            bg=self.colors['white']
        )
        
        self.next_phone_btn = tk.Button(
            self.phone_nav_frame,
            text="►",
            command=self.next_phone,
            font=('Arial', 8),
            bg=self.colors['secondary'],
            fg='white',
            padx=5,
            pady=1,
            cursor='hand2'
        )
        
        # Addresses - scrollable with max height
        tk.Label(self.result_content, text="Addresses:", font=('Arial', 8, 'bold'), 
                bg=self.colors['white']).grid(row=2, column=0, columnspan=2, sticky='w', padx=5, pady=(4, 2))
        
        # Create scrollable frame for addresses with max height
        addresses_container = tk.Frame(self.result_content, bg=self.colors['white'], height=150)
        addresses_container.grid(row=3, column=0, columnspan=2, sticky='ew', padx=5, pady=(0, 2))
        addresses_container.grid_propagate(False)  # Fixed max height
        
        addresses_canvas = tk.Canvas(addresses_container, bg=self.colors['white'], highlightthickness=0, height=150)
        addresses_scrollbar = ttk.Scrollbar(addresses_container, orient="vertical", command=addresses_canvas.yview)
        
        self.addresses_frame = tk.Frame(addresses_canvas, bg=self.colors['white'])
        addresses_canvas.create_window((0, 0), window=self.addresses_frame, anchor="nw")
        addresses_canvas.configure(yscrollcommand=addresses_scrollbar.set)
        
        addresses_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        addresses_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Enable mouse wheel scrolling for addresses
        def on_addresses_mousewheel(event):
            addresses_canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        addresses_canvas.bind("<Enter>", lambda e: addresses_canvas.bind_all("<MouseWheel>", on_addresses_mousewheel))
        addresses_canvas.bind("<Leave>", lambda e: addresses_canvas.unbind_all("<MouseWheel>"))
        
        # Update scroll region when addresses change
        self.addresses_canvas = addresses_canvas
        self.addresses_frame.bind('<Configure>', lambda e: addresses_canvas.configure(scrollregion=addresses_canvas.bbox("all")))
        
        # Fixed Status and Notes section (row 4 - always visible)
        status_notes_frame = tk.Frame(self.result_content, bg=self.colors['white'], height=60)
        status_notes_frame.grid(row=4, column=0, columnspan=2, sticky='ew', padx=5, pady=5)
        status_notes_frame.grid_propagate(False)  # Fixed height
        
        # Status
        tk.Label(status_notes_frame, text="Status:", font=('Arial', 9, 'bold'), 
                bg=self.colors['white']).pack(side=tk.LEFT, padx=(0, 5))
        
        self.status_var = tk.StringVar()
        self.status_dropdown = ttk.Combobox(
            status_notes_frame,
            textvariable=self.status_var,
            values=['', 'No answer', 'Not interested', 'Wrong Number', 'Call back', 'Appointment', 'DNC'],
            font=('Arial', 10),
            width=18,
            state='readonly'
        )
        self.status_dropdown.pack(side=tk.LEFT, padx=(0, 15))
        self.status_dropdown.bind('<<ComboboxSelected>>', self.on_status_change)
        
        # Notes
        tk.Label(status_notes_frame, text="Notes:", font=('Arial', 9, 'bold'), 
                bg=self.colors['white']).pack(side=tk.LEFT, padx=(0, 5))
        
        self.notes_text = tk.Text(
            status_notes_frame,
            font=('Arial', 9),
            width=30,
            height=2,
            wrap=tk.WORD,
            relief=tk.SOLID,
            borderwidth=1
        )
        self.notes_text.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        self.notes_text.bind('<FocusOut>', self.on_notes_change)
        self.notes_text.bind('<KeyRelease>', self.on_notes_typing)
        
        # Save Note button
        tk.Button(
            status_notes_frame,
            text="💾",
            command=self.save_note_with_status,
            font=('Arial', 10),
            bg=self.colors['success'],
            fg='white',
            padx=6,
            pady=4,
            cursor='hand2'
        ).pack(side=tk.LEFT)
        
        # Fixed navigation and manual buttons (row 5 - always visible)
        result_nav = tk.Frame(self.result_content, bg=self.colors['white'], height=40)
        result_nav.grid(row=5, column=0, columnspan=2, sticky='ew', pady=5)
        result_nav.grid_propagate(False)  # Fixed height
        
        # Center the buttons
        button_container = tk.Frame(result_nav, bg=self.colors['white'])
        button_container.pack(expand=True)
        
        tk.Button(
            button_container,
            text="◄ Prev",
            command=self.prev_result,
            font=('Arial', 9, 'bold'),
            bg=self.colors['secondary'],
            fg='white',
            padx=10,
            pady=4,
            cursor='hand2'
        ).pack(side=tk.LEFT, padx=3)
        
        tk.Button(
            button_container,
            text="Next ►",
            command=self.next_result,
            font=('Arial', 9, 'bold'),
            bg=self.colors['secondary'],
            fg='white',
            padx=10,
            pady=4,
            cursor='hand2'
        ).pack(side=tk.LEFT, padx=3)
        
        # Manual API call buttons
        tk.Button(
            button_container,
            text="📞 Phone",
            command=self.manual_phone_lookup,
            font=('Arial', 9),
            bg=self.colors['warning'],
            fg='white',
            padx=8,
            pady=4,
            cursor='hand2'
        ).pack(side=tk.LEFT, padx=3)
        
        tk.Button(
            button_container,
            text="📍 Address",
            command=self.manual_address_lookup,
            font=('Arial', 9),
            bg=self.colors['warning'],
            fg='white',
            padx=6,
            pady=3,
            cursor='hand2'
        ).pack(side=tk.LEFT, padx=2)
        
        # Person Navigation Section
        nav_frame = tk.Frame(main, bg=self.colors['primary'], height=35)
        nav_frame.grid(row=4, column=0, sticky='ew', padx=5, pady=(2, 3))
        nav_frame.grid_propagate(False)
        
        nav_content = tk.Frame(nav_frame, bg=self.colors['primary'])
        nav_content.pack(expand=True)
        
        tk.Button(
            nav_content,
            text="◄◄ Previous",
            command=self.prev_person,
            font=('Arial', 9, 'bold'),
            bg=self.colors['white'],
            fg=self.colors['primary'],
            padx=12,
            pady=5,
            cursor='hand2'
        ).pack(side=tk.LEFT, padx=5)
        
        self.person_counter_label = tk.Label(
            nav_content,
            text="Person 0 of 0",
            font=('Arial', 9, 'bold'),
            bg=self.colors['primary'],
            fg='white'
        )
        self.person_counter_label.pack(side=tk.LEFT, padx=15)
        
        tk.Button(
            nav_content,
            text="Next ►►",
            command=self.next_person,
            font=('Arial', 9, 'bold'),
            bg=self.colors['white'],
            fg=self.colors['primary'],
            padx=12,
            pady=5,
            cursor='hand2'
        ).pack(side=tk.LEFT, padx=5)
        
        # Settings Panel (right side)
        self.settings_panel = tk.LabelFrame(
            main,
            text="⚙️ SETTINGS",
            font=('Arial', 9, 'bold'),
            bg=self.colors['white'],
            fg=self.colors['primary'],
            relief=tk.RAISED,
            bd=1
        )
        self.settings_panel.grid(row=2, column=1, rowspan=3, sticky='nsew', padx=(0, 5), pady=2)
        
        # Settings content with scrollbar
        settings_canvas = tk.Canvas(self.settings_panel, bg=self.colors['white'], highlightthickness=0)
        settings_scrollbar = ttk.Scrollbar(self.settings_panel, orient="vertical", command=settings_canvas.yview)
        settings_content = tk.Frame(settings_canvas, bg=self.colors['white'])
        
        settings_canvas.create_window((0, 0), window=settings_content, anchor="nw")
        settings_canvas.configure(yscrollcommand=settings_scrollbar.set)
        
        settings_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        settings_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Enable mouse wheel scrolling for settings
        def on_settings_mousewheel(event):
            settings_canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        settings_canvas.bind("<Enter>", lambda e: settings_canvas.bind_all("<MouseWheel>", on_settings_mousewheel))
        settings_canvas.bind("<Leave>", lambda e: settings_canvas.unbind_all("<MouseWheel>"))
        
        # Initialize setting variables
        self.setting_vars = {}
        
        # API Settings Section
        tk.Label(
            settings_content,
            text="API Settings",
            font=('Arial', 9, 'bold'),
            bg=self.colors['white'],
            fg=self.colors['secondary']
        ).pack(pady=(5, 3), anchor='w', padx=5)
        
        # Only two checkboxes
        self.setting_vars['auto_phone_lookup'] = tk.BooleanVar(value=self.settings.get('auto_phone_lookup', True))
        phone_check = tk.Checkbutton(
            settings_content,
            text="Auto Phone Lookup",
            variable=self.setting_vars['auto_phone_lookup'],
            command=self.save_settings_auto,
            font=('Arial', 8),
            bg=self.colors['white'],
            activebackground=self.colors['white']
        )
        phone_check.pack(anchor='w', padx=10, pady=2)
        
        self.setting_vars['auto_address_lookup'] = tk.BooleanVar(value=self.settings.get('auto_address_lookup', True))
        address_check = tk.Checkbutton(
            settings_content,
            text="Auto Address Lookup",
            variable=self.setting_vars['auto_address_lookup'],
            command=self.save_settings_auto,
            font=('Arial', 8),
            bg=self.colors['white'],
            activebackground=self.colors['white']
        )
        address_check.pack(anchor='w', padx=10, pady=2)
        
        ttk.Separator(settings_content, orient='horizontal').pack(fill='x', pady=5)
        
        # Call History Settings
        tk.Label(
            settings_content,
            text="Call History",
            font=('Arial', 9, 'bold'),
            bg=self.colors['white'],
            fg=self.colors['secondary']
        ).pack(pady=(3, 3), anchor='w', padx=5)
        
        self.setting_vars['enable_call_history'] = tk.BooleanVar(value=self.settings.get('enable_call_history', True))
        tk.Checkbutton(
            settings_content,
            text="Track Call History",
            variable=self.setting_vars['enable_call_history'],
            command=self.save_settings_auto,
            font=('Arial', 8),
            bg=self.colors['white'],
            activebackground=self.colors['white']
        ).pack(anchor='w', padx=10, pady=2)
        
        # View Call History button
        tk.Button(
            settings_content,
            text="📋 View Call History",
            command=self.show_call_history_window,
            font=('Arial', 8),
            bg=self.colors['secondary'],
            fg='white',
            padx=8,
            pady=4,
            cursor='hand2'
        ).pack(anchor='w', padx=10, pady=5)
        
        # Update scroll region
        settings_content.update_idletasks()
        settings_canvas.config(scrollregion=settings_canvas.bbox("all"))
        
        # Update status bar based on settings
        self.update_settings_status()
    
    def toggle_settings_panel(self):
        """Toggle settings panel visibility"""
        if self.settings_visible.get():
            self.settings_panel.grid()
        else:
            self.settings_panel.grid_remove()
    
    def on_api_setting_change(self):
        """Handle main API checkbox change - uncheck children if parent unchecked"""
        if not self.setting_vars['auto_api_calls'].get():
            # Uncheck child checkboxes
            self.setting_vars['auto_phone_lookup'].set(False)
            self.setting_vars['auto_address_lookup'].set(False)
        self.save_settings_auto()
    
    def save_settings_auto(self):
        """Auto-save settings when changed"""
        for key, var in self.setting_vars.items():
            self.settings[key] = var.get()
        
        # Save to file
        try:
            with open('dialer_settings.json', 'w') as f:
                json.dump(self.settings, f, indent=2)
        except:
            pass
        
        # Update status bar
        self.update_settings_status()
        self.update_status("Settings saved", self.colors['success'])
        self.root.after(2000, lambda: self.update_status("Ready"))
    
    def update_settings_status(self):
        """Update status bar to reflect current settings"""
        # API status
        if self.settings.get('auto_api_calls', True):
            self.status_api.config(text="API: Auto", fg=self.colors['success'])
        else:
            self.status_api.config(text="API: Manual", fg=self.colors['warning'])
        
        # Cache status
        if self.settings.get('enable_cache', True):
            self.status_cache.config(text="Cache: On", fg=self.colors['success'])
        else:
            self.status_cache.config(text="Cache: Off", fg=self.colors['danger'])
    
    def update_status(self, message: str, color: str = None):
        """Update status bar"""
        self.status_bar.config(text=message)
        if color:
            self.status_bar.config(bg=color, fg='white' if color != self.colors['light'] else self.colors['primary'])
        else:
            self.status_bar.config(bg=self.colors['light'], fg=self.colors['primary'])
        self.root.update()
    
    def load_person(self, index):
        """Load person at given index"""
        if index < 0 or index >= len(self.original_data):
            return
        
        self.current_person_idx = index
        person = self.original_data[index]
        
        # Clear AI results for new person
        self.ai_results = None
        self.clear_ai_tab()
        
        # Start preloading next person in background (with proper validation)
        self.start_preloading_next_person()
        
        # Update counter
        self.person_counter_label.config(text=f"Person {index + 1} of {len(self.original_data)}")
        
        # Display original data
        self.orig_labels['Original Name'].config(state='normal')
        self.orig_labels['Original Name'].delete(0, tk.END)
        self.orig_labels['Original Name'].insert(0, person.get('name', ''))
        self.orig_labels['Original Name'].config(state='readonly')
        
        self.orig_labels['Original Phone'].config(state='normal')
        self.orig_labels['Original Phone'].delete(0, tk.END)
        self.orig_labels['Original Phone'].insert(0, person.get('phone', ''))
        self.orig_labels['Original Phone'].config(state='readonly')
        
        self.orig_labels['Original Address'].config(state='normal')
        self.orig_labels['Original Address'].delete(0, tk.END)
        self.orig_labels['Original Address'].insert(0, person.get('address', ''))
        self.orig_labels['Original Address'].config(state='readonly')
        
        self.orig_labels['Age'].config(state='normal')
        self.orig_labels['Age'].delete(0, tk.END)
        self.orig_labels['Age'].insert(0, str(person.get('age', '')))
        self.orig_labels['Age'].config(state='readonly')
        
        # Check permanent cache for raw API responses
        original_phone = self.lead_processor.clean_phone(person.get('phone', ''))
        
        # Check if we have preloaded AI results for this person
        preloaded_ai = None
        if hasattr(self, 'preloaded_ai_results') and original_phone in self.preloaded_ai_results:
            preloaded_ai = self.preloaded_ai_results[original_phone]
            del self.preloaded_ai_results[original_phone]  # Remove after use
            logger.info(f"Using preloaded AI analysis for {person.get('name', 'Unknown')}")
        
        # Try to load from permanent cache (raw API responses and AI analysis)
        if self.lead_processor.cache_manager:
            cached_lookup = self.lead_processor.cache_manager.get_cached_lookup(original_phone)
            if cached_lookup:
                phone_data, address_data, cached_ai_analysis = cached_lookup
                # Check if cached data has actual results
                phone_has_data = phone_data and phone_data.get('owners')
                address_has_data = address_data and address_data.get('current_residents')
                
                if phone_has_data or address_has_data:
                    # Process cached API responses as if they came from API
                    self.update_status("Loading data from cache...", self.colors['secondary'])
                    results = self.process_lookup_data(person, phone_data, address_data)
                    self.current_results = results
                    self.current_result_idx = 0
                    self.current_phone_idx = 0
                    
                    # AI tab is now default (tab 0), Standard View is tab 1
                    # Keep AI tab selected by default
                    self.display_current_result()
                    
                    # Handle AI analysis - use cached if available, otherwise run new analysis
                    if self.settings.get('ai_enabled') and self.settings.get('ai_person_filtering', True):
                        # Check if we have any meaningful results from accumulated data
                        has_results = False
                        if results and len(results) > 0:
                            # Check if we have actual results (not just "NO RESULTS FOUND")
                            for result in results:
                                if result.get('new_name') != 'NO RESULTS FOUND':
                                    has_results = True
                                    break
                        
                        if has_results:
                            # Priority: preloaded > cached > new analysis
                            if preloaded_ai:
                                # Use preloaded AI (from background preloading)
                                self.ai_results = preloaded_ai
                                self.show_ai_tab(self.ai_tab)
                                self.update_status("Data loaded instantly (preloaded AI) - no wait time!", self.colors['success'])
                            elif cached_ai_analysis:
                                # Use cached AI analysis
                                self.ai_results = cached_ai_analysis
                                self.show_ai_tab(self.ai_tab)
                                self.update_status("Data loaded from cache (including AI analysis) - no API cost", self.colors['success'])
                            elif phone_data or address_data:
                                # Run new AI analysis in background only if not cached or preloaded
                                threading.Thread(target=self._run_ai_analysis_background, 
                                               args=(person, phone_data, address_data, original_phone, index), 
                                               daemon=True).start()
                                self.update_status("Data loaded from cache - running AI analysis...", self.colors['secondary'])
                    
                    # Save to Excel and update Excel cache
                    self.save_results_to_excel(results)
                    if results:
                        self.excel_cache.update(original_phone, results[0])
                    
                    if not cached_ai_analysis:
                        self.update_status("Data loaded from cache - no API cost", self.colors['success'])
                    return
        
        # No cache or empty cache - perform API lookups
        self.update_status("Loading data from API...", self.colors['warning'])
        self.root.update()
        threading.Thread(target=self.perform_lookups, args=(person,), daemon=True).start()
    
    def perform_lookups(self, person, force_refresh=False):
        """Perform API lookups with cache support and auto settings"""
        try:
            original_phone = self.lead_processor.clean_phone(person.get('phone', ''))
            cache_manager = self.lead_processor.cache_manager
            was_cached = False
            cached_ai_analysis = None
            
            # Check cache first (unless force refresh)
            phone_data = None
            address_data = None
            
            if not force_refresh and cache_manager:
                cached_lookup = cache_manager.get_cached_lookup(original_phone)
                if cached_lookup:
                    phone_data, address_data, cached_ai_analysis = cached_lookup
                    # Check if cached data is empty - if so, try API instead
                    phone_has_data = phone_data and phone_data.get('owners')
                    address_has_data = address_data and address_data.get('current_residents')
                    
                    if phone_has_data or address_has_data:
                        was_cached = True
                    else:
                        # Cached but empty - treat as not cached to trigger API call
                        phone_data = None
                        address_data = None
                        cached_ai_analysis = None
            
            # Determine what to fetch based on auto settings
            auto_phone = self.settings.get('auto_phone_lookup', True)
            auto_address = self.settings.get('auto_address_lookup', True)
            
            # If not cached or force refresh, make API calls based on settings
            if not was_cached or force_refresh:
                # Phone lookup
                if auto_phone:
                    phone_data = self.lead_processor.phone_lookup(original_phone)
                
                # Address lookup with AI correction
                if auto_address:
                    street = person.get('address', '')
                    city = person.get('city', '')
                    state = person.get('state', '')
                    zip_code = person.get('zip', '')
                    
                    if street and (not city or not state or not zip_code):
                        parsed = self.lead_processor.parse_address(street)
                        street = parsed['street'] or street
                        city = parsed['city'] or city
                        state = parsed['state'] or state
                        zip_code = parsed['zip'] or zip_code
                    
                    if any([street, city, state, zip_code]):
                        # Enable AI correction if AI is enabled
                        enable_ai = self.settings.get('ai_enabled') and self.settings.get('ai_address_correction', True)
                        
                        # Pass status callback for AI updates
                        def status_update(msg):
                            self.root.after(0, lambda: self.update_status(msg, self.colors['secondary']))
                        
                        address_data, correction_info = self.lead_processor.address_lookup(
                            street, city, state, zip_code,
                            enable_ai_correction=enable_ai,
                            status_callback=status_update if enable_ai else None
                        )
                        
                        # Log AI correction if it happened
                        if correction_info:
                            self.ai_correction_log.append({
                                'person_idx': self.current_person_idx,
                                'original_address': f"{street}, {city}, {state} {zip_code}",
                                'corrected_address': f"{correction_info['corrected'].get('street_line_1', '')}, {correction_info['corrected'].get('city', '')}, {correction_info['corrected'].get('state_code', '')} {correction_info['corrected'].get('postal_code', '')}",
                                'reasoning': correction_info['reasoning']
                            })
                
                # Store in cache only if we have actual results
                if cache_manager:
                    # Only store if phone_data has owners or address_data has residents
                    phone_has_results = phone_data and phone_data.get('owners')
                    address_has_results = address_data and address_data.get('current_residents')
                    
                    if phone_has_results or address_has_results:
                        cache_manager.store_lookup(original_phone, phone_data or {}, address_data or {})
            
            # Process the data into results
            results = self.process_lookup_data(person, phone_data, address_data)
            
            self.current_results = results
            self.current_result_idx = 0
            self.current_phone_idx = 0
            
            # Save to Excel and update Excel cache immediately
            self.save_results_to_excel(results)
            if results:
                self.excel_cache.update(original_phone, results[0])
            
            # AI tab is now default (tab 0) - keep it selected
            self.root.after(0, self.display_current_result)
            
            # Run AI analysis if enabled and we have accumulated data
            if self.settings.get('ai_enabled') and self.settings.get('ai_person_filtering', True):
                # Check if we have any meaningful results from accumulated data
                has_results = False
                if results and len(results) > 0:
                    # Check if we have actual results (not just "NO RESULTS FOUND")
                    for result in results:
                        if result.get('new_name') != 'NO RESULTS FOUND':
                            has_results = True
                            break
                
                if has_results and (phone_data or address_data):
                    # Use cached AI analysis if available
                    if cached_ai_analysis:
                        self.ai_results = cached_ai_analysis
                        self.root.after(0, lambda: self.show_ai_tab(self.ai_tab))
                        self.root.after(0, lambda: self.update_status("Data loaded from cache (including AI analysis) - no API cost", self.colors['success']))
                    else:
                        # Run AI analysis in background only if not cached
                        # Get person index from current_person_idx
                        person_idx = self.current_person_idx
                        threading.Thread(target=self._run_ai_analysis_background, 
                                       args=(person, phone_data, address_data, original_phone, person_idx), 
                                       daemon=True).start()
            
            # Show appropriate status message
            if force_refresh:
                self.root.after(0, lambda: self.update_status("Data refreshed from API - cache updated", self.colors['success']))
            elif was_cached:
                self.root.after(0, lambda: self.update_status("Data loaded from cache - no API cost", self.colors['success']))
            elif not auto_phone and not auto_address:
                self.root.after(0, lambda: self.update_status("Auto lookups disabled - use manual buttons to fetch data", self.colors['warning']))
            else:
                self.root.after(0, lambda: self.update_status("Data loaded from API - cached for future use"))
        except Exception as e:
            error_msg = f"Error: {str(e)}"
            self.root.after(0, lambda: self.update_status(error_msg, self.colors['danger']))
    
    def extract_people_from_data(self, data):
        """Extract all people, phones, and addresses from API response using recursive traversal"""
        people = []
        
        def traverse(obj, path=""):
            """Recursively traverse the data structure to find people"""
            if not obj or not isinstance(obj, dict):
                return
            
            # Check if this object represents a person
            if obj.get('name') or obj.get('firstname') or obj.get('lastname'):
                person_name = obj.get('name') or f"{obj.get('firstname', '')} {obj.get('lastname', '')}".strip()
                
                if person_name:
                    person_info = {
                        'name': person_name,
                        'age': self.lead_processor.extract_age(obj),
                        'phones': [],
                        'addresses': []
                    }
                    
                    # Extract phones from alternate_phones (reverse phone lookup)
                    if obj.get('alternate_phones') and isinstance(obj['alternate_phones'], list):
                        for phone_obj in obj['alternate_phones']:
                            if isinstance(phone_obj, dict) and phone_obj.get('phoneNumber'):
                                if phone_obj['phoneNumber'] not in person_info['phones']:
                                    person_info['phones'].append(phone_obj['phoneNumber'])
                    
                    # Extract phones from phones array (reverse address lookup)
                    if obj.get('phones') and isinstance(obj['phones'], list):
                        for phone_obj in obj['phones']:
                            phone_num = None
                            if isinstance(phone_obj, dict):
                                phone_num = phone_obj.get('phone_number') or phone_obj.get('phoneNumber') or phone_obj.get('phone')
                            elif isinstance(phone_obj, str):
                                phone_num = phone_obj
                            
                            if phone_num and phone_num not in person_info['phones']:
                                person_info['phones'].append(phone_num)
                    
                    # Extract from phone_numbers array
                    if obj.get('phone_numbers') and isinstance(obj['phone_numbers'], list):
                        for phone_obj in obj['phone_numbers']:
                            phone_num = None
                            if isinstance(phone_obj, dict):
                                phone_num = phone_obj.get('phone_number') or phone_obj.get('phoneNumber')
                            elif isinstance(phone_obj, str):
                                phone_num = phone_obj
                            
                            if phone_num and phone_num not in person_info['phones']:
                                person_info['phones'].append(phone_num)
                    
                    # Single phone field
                    if obj.get('phone') and obj['phone'] not in person_info['phones']:
                        person_info['phones'].append(obj['phone'])
                    
                    # Extract addresses from current_addresses
                    if obj.get('current_addresses') and isinstance(obj['current_addresses'], list):
                        for addr in obj['current_addresses']:
                            if isinstance(addr, dict):
                                formatted_addr = self.lead_processor.format_address(addr)
                                if formatted_addr and formatted_addr not in person_info['addresses']:
                                    person_info['addresses'].append(formatted_addr)
                    
                    # Extract addresses from historical_addresses
                    if obj.get('historical_addresses') and isinstance(obj['historical_addresses'], list):
                        for addr in obj['historical_addresses']:
                            if isinstance(addr, dict):
                                formatted_addr = self.lead_processor.format_address(addr)
                                if formatted_addr and formatted_addr not in person_info['addresses']:
                                    person_info['addresses'].append(formatted_addr)
                    
                    people.append(person_info)
            
            # Continue traversing all nested objects and arrays
            for key, value in obj.items():
                if isinstance(value, dict):
                    traverse(value, f"{path}.{key}")
                elif isinstance(value, list):
                    for item in value:
                        if isinstance(item, dict):
                            traverse(item, f"{path}.{key}[]")
        
        # Handle both dict and list at top level
        if isinstance(data, dict):
            traverse(data)
        elif isinstance(data, list):
            for item in data:
                if isinstance(item, dict):
                    traverse(item)
        
        return people
    
    def process_lookup_data(self, person, phone_data, address_data):
        """Process phone and address data into result format using comprehensive extraction"""
        original_name = person.get('name', '')
        original_phone = self.lead_processor.clean_phone(person.get('phone', ''))
        original_address = person.get('address', '')
        
        found_people = []
        address_lookup_failed = False
        
        # Check if address lookup failed
        if not address_data or 'current_residents' not in address_data or not address_data['current_residents']:
            address_lookup_failed = True
        
        # Extract from phone data using comprehensive extraction
        if phone_data:
            phone_people = self.extract_people_from_data(phone_data)
            for person_info in phone_people:
                # Check if person already exists
                existing = next((p for p in found_people if p['name'] == person_info['name']), None)
                if existing:
                    # Merge phones and addresses
                    for phone in person_info['phones']:
                        if phone not in existing['phones']:
                            existing['phones'].append(phone)
                    for addr in person_info['addresses']:
                        if addr not in existing['addresses']:
                            existing['addresses'].append(addr)
                else:
                    found_people.append(person_info)
        
        # Extract from address data using comprehensive extraction
        if address_data:
            address_people = self.extract_people_from_data(address_data)
            for person_info in address_people:
                # Check if person already exists
                existing = next((p for p in found_people if p['name'] == person_info['name']), None)
                if existing:
                    # Merge phones and addresses
                    for phone in person_info['phones']:
                        if phone not in existing['phones']:
                            existing['phones'].append(phone)
                    for addr in person_info['addresses']:
                        if addr not in existing['addresses']:
                            existing['addresses'].append(addr)
                else:
                    found_people.append(person_info)
        
        # Create results
        results = []
        if not found_people:
            results.append({
                'original_address': original_address,
                'original_phone': original_phone,
                'blank': '',
                'age': '',
                'original_name': original_name,
                'new_phones': [],
                'new_addresses': [],
                'new_name': 'NO RESULTS FOUND',
                'status': '',
                'notes': '',
                'address_lookup_failed': address_lookup_failed
            })
        else:
            for person_info in found_people:
                results.append({
                    'original_address': original_address,
                    'original_phone': original_phone,
                    'blank': '',
                    'age': person_info['age'],
                    'original_name': original_name,
                    'new_phones': person_info['phones'],
                    'new_addresses': person_info['addresses'],
                    'new_name': person_info['name'],
                    'status': '',
                    'notes': '',
                    'address_lookup_failed': address_lookup_failed
                })
        
        return results
    
    def accumulate_phone_results(self, phone_data, original_person, trigger_ai=True):
        """Accumulate phone lookup results with existing data using comprehensive extraction"""
        if not self.current_results or self.current_results[0].get('new_name') == 'NO RESULTS FOUND':
            # No existing results, create new ones
            results = self.process_lookup_data(original_person, phone_data, None)
            self.current_results = results
        else:
            # Extract people from phone data
            phone_people = self.extract_people_from_data(phone_data) if phone_data else []
            
            for person_info in phone_people:
                # Find existing person or create new
                existing_result = next((r for r in self.current_results if r['new_name'] == person_info['name']), None)
                
                if existing_result:
                    # Merge phones and addresses
                    for phone in person_info['phones']:
                        if phone not in existing_result['new_phones']:
                            existing_result['new_phones'].append(phone)
                    for addr in person_info['addresses']:
                        if addr not in existing_result['new_addresses']:
                            existing_result['new_addresses'].append(addr)
                    # Update age if we have it
                    if person_info['age'] and not existing_result.get('age'):
                        existing_result['age'] = person_info['age']
                else:
                    # Create new person entry
                    new_result = {
                        'original_address': original_person.get('address', ''),
                        'original_phone': self.lead_processor.clean_phone(original_person.get('phone', '')),
                        'blank': '',
                        'age': person_info['age'],
                        'original_name': original_person.get('name', ''),
                        'new_phones': person_info['phones'],
                        'new_addresses': person_info['addresses'],
                        'new_name': person_info['name'],
                        'status': '',
                        'notes': '',
                        'address_lookup_failed': False
                    }
                    self.current_results.append(new_result)
        
        # Save and display
        original_phone = self.lead_processor.clean_phone(original_person.get('phone', ''))
        self.save_results_to_excel(self.current_results)
        if self.current_results:
            self.excel_cache.update(original_phone, self.current_results[0])
        
        # Trigger AI analysis if enabled and we have accumulated data
        if trigger_ai and self.settings.get('ai_enabled') and self.settings.get('ai_person_filtering', True):
            # Get both phone and address data from cache for AI analysis
            phone_api_data = phone_data
            address_api_data = {}
            if self.lead_processor.cache_manager:
                cached = self.lead_processor.cache_manager.get_cached_lookup(original_phone)
                if cached:
                    address_api_data = cached[1] or {}
            
            # Check if we have any meaningful results from accumulated data
            has_results = False
            if self.current_results and len(self.current_results) > 0:
                for result in self.current_results:
                    if result.get('new_name') != 'NO RESULTS FOUND':
                        has_results = True
                        break
            
            # Run AI analysis if we have any accumulated data with results
            if has_results and (phone_api_data or address_api_data):
                self.ai_results = self.process_ai_analysis(original_person, phone_api_data, address_api_data)
                # Update AI tab
                self.show_ai_tab(self.ai_tab)
        
        # AI tab is now default (tab 0) - keep it selected
        
        self.display_current_result()
    
    def accumulate_address_results(self, address_data, original_person, trigger_ai=True):
        """Accumulate address lookup results with existing data using comprehensive extraction"""
        if not self.current_results or self.current_results[0].get('new_name') == 'NO RESULTS FOUND':
            # No existing results, create new ones
            results = self.process_lookup_data(original_person, None, address_data)
            self.current_results = results
        else:
            # Extract people from address data
            address_people = self.extract_people_from_data(address_data) if address_data else []
            
            for person_info in address_people:
                # Find existing person or create new
                existing_result = next((r for r in self.current_results if r['new_name'] == person_info['name']), None)
                
                if existing_result:
                    # Merge phones and addresses
                    for phone in person_info['phones']:
                        if phone not in existing_result['new_phones']:
                            existing_result['new_phones'].append(phone)
                    for addr in person_info['addresses']:
                        if addr not in existing_result['new_addresses']:
                            existing_result['new_addresses'].append(addr)
                    # Update age if we have it
                    if person_info['age'] and not existing_result.get('age'):
                        existing_result['age'] = person_info['age']
                    # Clear address lookup failed flag
                    existing_result['address_lookup_failed'] = False
                else:
                    # Create new person entry
                    new_result = {
                        'original_address': original_person.get('address', ''),
                        'original_phone': self.lead_processor.clean_phone(original_person.get('phone', '')),
                        'blank': '',
                        'age': person_info['age'],
                        'original_name': original_person.get('name', ''),
                        'new_phones': person_info['phones'],
                        'new_addresses': person_info['addresses'],
                        'new_name': person_info['name'],
                        'status': '',
                        'notes': '',
                        'address_lookup_failed': False
                    }
                    self.current_results.append(new_result)
        
        # Save and display
        original_phone = self.lead_processor.clean_phone(original_person.get('phone', ''))
        self.save_results_to_excel(self.current_results)
        if self.current_results:
            self.excel_cache.update(original_phone, self.current_results[0])
        
        # Trigger AI analysis if enabled and we have accumulated data
        if trigger_ai and self.settings.get('ai_enabled') and self.settings.get('ai_person_filtering', True):
            # Get both phone and address data from cache for AI analysis
            address_api_data = address_data
            phone_api_data = {}
            if self.lead_processor.cache_manager:
                cached = self.lead_processor.cache_manager.get_cached_lookup(original_phone)
                if cached:
                    phone_api_data = cached[0] or {}
            
            # Check if we have any meaningful results from accumulated data
            has_results = False
            if self.current_results and len(self.current_results) > 0:
                for result in self.current_results:
                    if result.get('new_name') != 'NO RESULTS FOUND':
                        has_results = True
                        break
            
            # Run AI analysis if we have any accumulated data with results
            if has_results and (phone_api_data or address_api_data):
                self.ai_results = self.process_ai_analysis(original_person, phone_api_data, address_api_data)
                # Update AI tab
                self.show_ai_tab(self.ai_tab)
        
        # AI tab is now default (tab 0) - keep it selected
        
        self.display_current_result()
    
    def _run_ai_analysis_background(self, person, phone_data, address_data, original_phone, person_idx):
        """Run AI analysis in background thread and update cache"""
        try:
            ai_results = self.process_ai_analysis(person, phone_data, address_data)
            if ai_results:
                # Update cache with AI analysis
                if self.lead_processor.cache_manager:
                    self.lead_processor.cache_manager.update_ai_analysis(original_phone, ai_results)
                
                # Update UI on main thread ONLY if still viewing this person
                self.root.after(0, lambda: self._update_ai_results(ai_results, original_phone, person_idx))
            else:
                self.root.after(0, lambda: self.update_status("AI analysis completed", self.colors['success']))
        except Exception as e:
            logger.error(f"Background AI analysis error: {e}")
            self.root.after(0, lambda: self.update_status("AI analysis error", self.colors['danger']))
    
    def _update_ai_results(self, ai_results, original_phone, person_idx):
        """Update AI results on main thread - ONLY if still viewing this person"""
        # Validate we're still on the same person
        if self.current_person_idx != person_idx:
            logger.info(f"Skipping AI update - user navigated away (was {person_idx}, now {self.current_person_idx})")
            return
        
        # Double-check the phone number matches current person
        if self.current_person_idx < len(self.original_data):
            current_person = self.original_data[self.current_person_idx]
            current_phone = self.lead_processor.clean_phone(current_person.get('phone', ''))
            if current_phone != original_phone:
                logger.info(f"Skipping AI update - phone mismatch")
                return
        
        # Safe to update UI
        self.ai_results = ai_results
        self.show_ai_tab(self.ai_tab)
        self.update_status("AI analysis complete", self.colors['success'])
    
    def clear_ai_tab(self):
        """Clear the AI tab content"""
        # Clear existing content
        for widget in self.ai_tab.winfo_children():
            widget.destroy()
        
        # Show "Loading..." or "No Analysis" message
        tk.Label(
            self.ai_tab,
            text="No AI Analysis Available",
            font=('Arial', 14, 'bold'),
            bg='white',
            fg='#999'
        ).pack(pady=50)
        
        tk.Label(
            self.ai_tab,
            text="AI analysis will appear here after API lookups complete",
            font=('Arial', 10),
            bg='white',
            fg='#666'
        ).pack()
    
    def start_preloading_next_person(self):
        """Start preloading the next person's data in background with proper validation"""
        next_idx = self.current_person_idx + 1
        if next_idx < len(self.original_data):
            # Start background thread to preload next person
            threading.Thread(
                target=self.preload_person_data,
                args=(next_idx,),
                daemon=True
            ).start()
    
    def preload_person_data(self, person_idx):
        """Preload person data in background thread using existing cache-checking logic"""
        try:
            if person_idx >= len(self.original_data):
                return
            
            person = self.original_data[person_idx]
            original_phone = self.lead_processor.clean_phone(person.get('phone', ''))
            
            # Use the same logic as perform_lookups but in background
            cache_manager = self.lead_processor.cache_manager
            phone_data = None
            address_data = None
            
            # Check cache first
            if cache_manager:
                cached_lookup = cache_manager.get_cached_lookup(original_phone)
                if cached_lookup:
                    phone_data, address_data, cached_ai_analysis = cached_lookup
                    # Check if cached data is empty - if so, try API instead
                    phone_has_data = phone_data and phone_data.get('owners')
                    address_has_data = address_data and address_data.get('current_residents')
                    
                    if phone_has_data or address_has_data:
                        # Already cached with good data, check if AI analysis is also cached
                        logger.info(f"Using cached data for preloading {person.get('name', 'Unknown')}")
                        if self.settings.get('ai_enabled') and self.settings.get('ai_person_filtering', True):
                            if cached_ai_analysis:
                                # AI analysis is also cached, store it for instant loading
                                if not hasattr(self, 'preloaded_ai_results'):
                                    self.preloaded_ai_results = {}
                                self.preloaded_ai_results[original_phone] = cached_ai_analysis
                                logger.info(f"Using cached AI analysis for preloading {person.get('name', 'Unknown')}")
                            else:
                                # Run AI analysis on cached data
                                self.preload_ai_analysis(person, phone_data, address_data, original_phone)
                        return
            
            # Not cached or empty cache, perform API lookups
            auto_phone = self.settings.get('auto_phone_lookup', True)
            auto_address = self.settings.get('auto_address_lookup', True)
            
            # Phone lookup
            if auto_phone:
                try:
                    phone_data = self.lead_processor.phone_lookup(original_phone)
                except Exception as e:
                    logger.error(f"Preload phone lookup error: {e}")
                    phone_data = None
            
            # Address lookup with AI correction support
            if auto_address:
                try:
                    street = person.get('address', '')
                    city = person.get('city', '')
                    state = person.get('state', '')
                    zip_code = person.get('zip', '')
                    
                    # Parse address if needed
                    if street and (not city or not state or not zip_code):
                        parsed = self.lead_processor.parse_address(street)
                        street = parsed['street'] or street
                        city = parsed['city'] or city
                        state = parsed['state'] or state
                        zip_code = parsed['zip'] or zip_code
                    
                    if any([street, city, state, zip_code]):
                        # Use address_lookup with AI correction enabled
                        enable_ai_correction = (self.settings.get('ai_enabled') and 
                                              self.settings.get('ai_address_correction', True) and 
                                              self.ai_assistant)
                        
                        result = self.lead_processor.address_lookup(
                            street, city, state, zip_code, 
                            enable_ai_correction=enable_ai_correction
                        )
                        
                        # Handle tuple return (result, correction_info)
                        if isinstance(result, tuple):
                            address_data, correction_info = result
                        else:
                            address_data = result
                            
                except Exception as e:
                    logger.error(f"Preload address lookup error: {e}")
                    address_data = None
            
            # Store in cache if we have results
            if cache_manager:
                phone_has_results = phone_data and phone_data.get('owners')
                address_has_results = address_data and address_data.get('current_residents')
                
                if phone_has_results or address_has_results:
                    cache_manager.store_lookup(original_phone, phone_data or {}, address_data or {})
                    logger.info(f"Cached preloaded data for {person.get('name', 'Unknown')}")
            
            # Preload AI analysis if we have meaningful data
            if self.settings.get('ai_enabled') and self.settings.get('ai_person_filtering', True):
                if phone_data or address_data:
                    self.preload_ai_analysis(person, phone_data, address_data, original_phone)
                
        except Exception as e:
            logger.error(f"Preload person data error: {e}")
    
    def preload_ai_analysis(self, person, phone_data, address_data, original_phone):
        """Preload AI analysis in background and store in both memory and permanent cache"""
        try:
            if not self.ai_assistant:
                return
            
            # Check if we have meaningful results
            results = self.process_lookup_data(person, phone_data, address_data)
            has_results = False
            if results and len(results) > 0:
                for result in results:
                    if result.get('new_name') != 'NO RESULTS FOUND':
                        has_results = True
                        break
            
            if has_results and (phone_data or address_data):
                ai_results = self.process_ai_analysis(person, phone_data, address_data)
                if ai_results:
                    # Store in permanent cache first
                    if self.lead_processor.cache_manager:
                        self.lead_processor.cache_manager.update_ai_analysis(original_phone, ai_results)
                        logger.info(f"Stored preloaded AI analysis in permanent cache for {person.get('name', 'Unknown')}")
                    
                    # Also store in memory for instant access
                    if not hasattr(self, 'preloaded_ai_results'):
                        self.preloaded_ai_results = {}
                    self.preloaded_ai_results[original_phone] = ai_results
                    logger.info(f"Preloaded AI analysis in memory for {person.get('name', 'Unknown')}")
                    
        except Exception as e:
            logger.error(f"Preload AI analysis error: {e}")
    
    def refresh_current_person(self):
        """Refresh current person data by forcing fresh API calls (no confirmation)"""
        if self.current_person_idx < 0 or self.current_person_idx >= len(self.original_data):
            return
        
        person = self.original_data[self.current_person_idx]
        
        # Disable refresh button during refresh
        self.refresh_btn.config(state='disabled', text='...')
        self.update_status(f"Refreshing {person.get('name', 'person')} from API...", self.colors['warning'])
        self.root.update()
        
        # Perform refresh in background thread
        def refresh_worker():
            try:
                self.perform_lookups(person, force_refresh=True)
                self.root.after(0, lambda: self.refresh_btn.config(state='normal', text='🔄'))
                self.root.after(0, lambda: self.update_status("Data refreshed from API", self.colors['success']))
                self.root.after(3000, lambda: self.update_status("Ready"))
            except Exception as e:
                error_msg = f"Refresh error: {str(e)}"
                self.root.after(0, lambda: self.refresh_btn.config(state='normal', text='🔄'))
                self.root.after(0, lambda: self.update_status(error_msg, self.colors['danger']))
        
        threading.Thread(target=refresh_worker, daemon=True).start()
    
    def save_results_to_excel(self, results, only_if_has_data=True):
        """Save results to output Excel - updates existing row or creates new one"""
        try:
            # Don't save if no real data (only "NO RESULTS FOUND")
            if only_if_has_data:
                has_real_data = False
                for result in results:
                    if result.get('new_name') != 'NO RESULTS FOUND':
                        if result.get('new_phones') or result.get('new_addresses'):
                            has_real_data = True
                            break
                
                if not has_real_data:
                    return  # Don't save empty results
            
            if os.path.exists(self.output_excel_file):
                wb = load_workbook(self.output_excel_file)
                ws = wb.active
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = "Lead Results"
                
                headers = ['Original Address', 'Original Phone', '', 'Age', 'Original Name', 
                          'New Phone', 'New Address', 'New Name', 'Status', 'Notes']
                ws.append(headers)
                
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = Font(bold=True, color="FFFFFF")
            
            for result in results:
                phones_str = ", ".join(result.get('new_phones', []))
                addresses_str = " | ".join(result.get('new_addresses', []))
                
                # Check if row already exists for this original phone
                original_phone = result['original_phone']
                existing_row = None
                
                for row_idx in range(2, ws.max_row + 1):
                    if ws.cell(row_idx, 2).value == original_phone:  # Column B is Original Phone
                        existing_row = row_idx
                        break
                
                row_data = [
                    result['original_address'],
                    result['original_phone'],
                    '',
                    result['age'],
                    result['original_name'],
                    phones_str,
                    addresses_str,
                    result['new_name'],
                    result.get('status', ''),
                    result.get('notes', '')
                ]
                
                if existing_row:
                    # Update existing row
                    for col_idx, value in enumerate(row_data, start=1):
                        ws.cell(existing_row, col_idx).value = value
                    current_row = existing_row
                else:
                    # Add new row
                    ws.append(row_data)
                    current_row = ws.max_row
                
                # Apply formatting
                if result.get('address_lookup_failed', False):
                    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    for cell in ws[current_row]:
                        cell.fill = red_fill
                        cell.font = Font(color="FFFFFF")
                else:
                    orig_last = self.lead_processor.get_last_name(result['original_name'])
                    new_last = self.lead_processor.get_last_name(result['new_name'])
                    if orig_last and new_last and orig_last == new_last:
                        green_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                        for cell in ws[current_row]:
                            cell.fill = green_fill
            
            status_options = "No answer,Not interested,Wrong Number,Call back,Appointment,DNC"
            status_dv = DataValidation(type="list", formula1=f'"{status_options}"', allow_blank=True)
            status_dv.add(f'I2:I{ws.max_row}')
            ws.add_data_validation(status_dv)
            
            wb.save(self.output_excel_file)
        except Exception as e:
            print(f"Error saving to Excel: {e}")
    
    def display_current_result(self):
        """Display current result - validates we're showing data for current person"""
        # Validate we have a valid person index
        if self.current_person_idx < 0 or self.current_person_idx >= len(self.original_data):
            logger.warning(f"Invalid person index: {self.current_person_idx}")
            return
        
        if not self.current_results:
            self.result_counter_label.config(text="No results found")
            self.new_name_entry.config(state='normal')
            self.new_name_entry.delete(0, tk.END)
            self.new_name_entry.config(state='readonly')
            self.age_entry.config(state='normal')
            self.age_entry.delete(0, tk.END)
            self.age_entry.config(state='readonly')
            self.phone_entry.config(state='normal')
            self.phone_entry.delete(0, tk.END)
            self.phone_entry.config(state='readonly')
            self.lastname_match_label.pack_forget()
            self.address_warning_label.pack_forget()
            self.status_var.set('')
            for widget in self.addresses_frame.winfo_children():
                widget.destroy()
            return
        
        result = self.current_results[self.current_result_idx]
        
        # Update counter
        self.result_counter_label.config(
            text=f"Result {self.current_result_idx + 1} of {len(self.current_results)} (Person: {result.get('new_name', 'Unknown')})"
        )
        
        # Update name
        self.new_name_entry.config(state='normal')
        self.new_name_entry.delete(0, tk.END)
        self.new_name_entry.insert(0, result.get('new_name', ''))
        self.new_name_entry.config(state='readonly')
        
        # Update age
        self.age_entry.config(state='normal')
        self.age_entry.delete(0, tk.END)
        self.age_entry.insert(0, str(result.get('age', '')))
        self.age_entry.config(state='readonly')
        
        # Display phone
        phones = result.get('new_phones', [])
        if phones:
            if self.current_phone_idx >= len(phones):
                self.current_phone_idx = 0
            current_phone = phones[self.current_phone_idx]
            self.phone_entry.config(state='normal')
            self.phone_entry.delete(0, tk.END)
            self.phone_entry.insert(0, current_phone)
            self.phone_entry.config(state='readonly')
            
            if len(phones) > 1:
                self.prev_phone_btn.pack(side=tk.LEFT, padx=2)
                self.phone_counter_label.config(text=f"{self.current_phone_idx + 1}/{len(phones)}")
                self.phone_counter_label.pack(side=tk.LEFT, padx=5)
                self.next_phone_btn.pack(side=tk.LEFT, padx=2)
            else:
                self.prev_phone_btn.pack_forget()
                self.phone_counter_label.pack_forget()
                self.next_phone_btn.pack_forget()
        else:
            self.phone_entry.config(state='normal')
            self.phone_entry.delete(0, tk.END)
            self.phone_entry.insert(0, "No phone found")
            self.phone_entry.config(state='readonly')
            self.prev_phone_btn.pack_forget()
            self.phone_counter_label.pack_forget()
            self.next_phone_btn.pack_forget()
        
        # Display addresses with individual buttons
        for widget in self.addresses_frame.winfo_children():
            widget.destroy()
        
        addresses = result.get('new_addresses', [])
        if addresses:
            for i, addr in enumerate(addresses):
                addr_row = tk.Frame(self.addresses_frame, bg=self.colors['white'])
                addr_row.pack(fill=tk.X, pady=2)
                
                addr_entry = tk.Entry(
                    addr_row,
                    font=('Arial', 9),
                    relief=tk.FLAT,
                    bg=self.colors['white'],
                    fg=self.colors['primary'],
                    readonlybackground=self.colors['white'],
                    state='readonly'
                )
                addr_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
                addr_entry.config(state='normal')
                addr_entry.insert(0, addr.strip())
                addr_entry.config(state='readonly')
                
                # Copy button for this address
                tk.Button(
                    addr_row,
                    text="📋",
                    command=lambda a=addr.strip(): self.copy_specific_address(a),
                    font=('Arial', 8),
                    bg=self.colors['light'],
                    padx=4,
                    pady=1,
                    cursor='hand2'
                ).pack(side=tk.LEFT, padx=1)
                
                # Google search button for this address
                tk.Button(
                    addr_row,
                    text="🔍",
                    command=lambda a=addr.strip(): self.google_search_specific_address(a),
                    font=('Arial', 8),
                    bg=self.colors['secondary'],
                    fg='white',
                    padx=4,
                    pady=1,
                    cursor='hand2'
                ).pack(side=tk.LEFT, padx=1)
        else:
            tk.Label(
                self.addresses_frame,
                text="No addresses found",
                font=('Arial', 9),
                bg=self.colors['white'],
                fg='gray'
            ).pack(anchor='w')
        
        # Address lookup failure warning
        if result.get('address_lookup_failed', False):
            self.address_warning_label.pack(side=tk.LEFT, padx=10)
        else:
            self.address_warning_label.pack_forget()
        
        # Last name match
        orig_last = self.lead_processor.get_last_name(result.get('original_name', ''))
        new_last = self.lead_processor.get_last_name(result.get('new_name', ''))
        
        if orig_last and new_last and orig_last == new_last:
            self.lastname_match_label.config(text=f"⚠️ LAST NAME MATCH: {orig_last}")
            self.lastname_match_label.pack(side=tk.RIGHT, padx=15)
        else:
            self.lastname_match_label.pack_forget()
        
        # Status
        self.status_var.set(result.get('status', ''))
        
        # Notes
        self.notes_text.delete('1.0', tk.END)
        self.notes_text.insert('1.0', result.get('notes', ''))
    
    def next_result(self):
        if self.current_result_idx < len(self.current_results) - 1:
            self.current_result_idx += 1
            self.current_phone_idx = 0
            self.display_current_result()
    
    def prev_result(self):
        if self.current_result_idx > 0:
            self.current_result_idx -= 1
            self.current_phone_idx = 0
            self.display_current_result()
    
    def next_phone(self):
        if not self.current_results:
            return
        result = self.current_results[self.current_result_idx]
        phones = result.get('new_phones', [])
        if self.current_phone_idx < len(phones) - 1:
            self.current_phone_idx += 1
            self.display_current_result()
    
    def prev_phone(self):
        if not self.current_results:
            return
        result = self.current_results[self.current_result_idx]
        phones = result.get('new_phones', [])
        if self.current_phone_idx > 0:
            self.current_phone_idx -= 1
            self.display_current_result()
    
    def next_person(self):
        # Save call history if call was made but not saved yet
        self.save_pending_call_history()
        
        if self.current_person_idx < len(self.original_data) - 1:
            self.load_person(self.current_person_idx + 1)
    
    def prev_person(self):
        # Save call history if call was made but not saved yet
        self.save_pending_call_history()
        
        if self.current_person_idx > 0:
            self.load_person(self.current_person_idx - 1)
    
    def save_pending_call_history(self):
        """Save call history if a call was made but user didn't set status"""
        if hasattr(self, 'last_called_phone') and self.last_called_phone:
            phone = self.last_called_phone
            name = self.last_called_name
            status = self.status_var.get() if hasattr(self, 'status_var') else ''
            notes = self.notes_text.get('1.0', tk.END).strip() if hasattr(self, 'notes_text') else ''
            
            # Save with whatever info we have
            if not status:
                status = 'Called'  # Default status if not set
            
            self.add_to_call_history(phone, name, status, notes)
            
            # Clear tracking
            self.last_called_phone = None
            self.last_called_name = None
    
    def copy_phone(self):
        phone = self.phone_entry.get()
        if phone and phone != "No phone found":
            self.root.clipboard_clear()
            self.root.clipboard_append(phone)
            self.update_status("Phone number copied to clipboard", self.colors['success'])
            self.root.after(2000, lambda: self.update_status("Ready"))
    
    def make_call(self):
        phone = self.phone_entry.get()
        
        if not phone or phone == "No phone found":
            self.update_status("No phone number to call", self.colors['warning'])
            return
        
        if not self.cloudtalk_api:
            self.copy_phone()
            self.update_status("CloudTalk not configured - phone number copied", self.colors['warning'])
            return
        
        # Flash the call button
        self.flash_call_button()
        
        self.call_btn.config(state='disabled', text='Calling...')
        self.update_status("Initiating call...", self.colors['warning'])
        self.root.update()
        
        # Get current person name for call history
        current_name = ""
        if self.current_results and self.current_result_idx < len(self.current_results):
            current_name = self.current_results[self.current_result_idx].get('name', '')
        
        threading.Thread(target=self.call_worker, args=(phone, current_name), daemon=True).start()
    
    def call_worker(self, phone, name):
        result = self.cloudtalk_api.make_call(self.agent_id, phone)
        
        self.root.after(0, lambda: self.call_btn.config(state='normal', text='📞 Call'))
        
        if result['success']:
            self.root.after(0, lambda: self.update_status("Call initiated successfully", self.colors['success']))
            # Mark that a call was just made (for call history tracking)
            self.last_called_phone = phone
            self.last_called_name = name
            # Flash status dropdown to remind user to set status
            self.root.after(0, self.flash_status_dropdown)
        else:
            self.root.after(0, lambda: self.update_status(result['message'], self.colors['danger']))
            self.root.after(0, self.copy_phone)
            # Add failed call to history immediately
            self.root.after(0, lambda: self.add_to_call_history(phone, name, "Call Failed", result['message']))
        
        self.root.after(3000, lambda: self.update_status("Ready"))
    
    def flash_status_dropdown(self):
        """Flash the status dropdown to remind user to set status"""
        original_style = ttk.Style()
        
        def flash(count=0):
            if count < 6:  # Flash 3 times (on/off = 2 counts per flash)
                if count % 2 == 0:
                    # Highlight
                    self.status_dropdown.configure(style='Flash.TCombobox')
                else:
                    # Normal
                    self.status_dropdown.configure(style='TCombobox')
                self.root.after(500, lambda: flash(count + 1))
        
        # Create flashing style
        style = ttk.Style()
        style.configure('Flash.TCombobox', 
                       fieldbackground='#FFF3CD',
                       background='#FFC107',
                       bordercolor='#FFC107',
                       borderwidth=2)
        
        flash()
    
    def copy_original_address(self):
        """Copy the original address to clipboard"""
        address = self.orig_labels['Original Address'].get()
        if address:
            self.root.clipboard_clear()
            self.root.clipboard_append(address)
            self.update_status("Original address copied to clipboard", self.colors['success'])
            self.root.after(2000, lambda: self.update_status("Ready"))
        else:
            self.update_status("No original address to copy", self.colors['warning'])
            self.root.after(2000, lambda: self.update_status("Ready"))
    
    def google_search_original_address(self):
        """Open Google search for the original address"""
        address = self.orig_labels['Original Address'].get()
        if address:
            search_url = f"https://www.google.com/search?q={urllib.parse.quote(address)}"
            webbrowser.open(search_url)
            self.update_status("Opening Google search for original address...", self.colors['success'])
            self.root.after(2000, lambda: self.update_status("Ready"))
        else:
            self.update_status("No original address to search", self.colors['warning'])
            self.root.after(2000, lambda: self.update_status("Ready"))
    
    def copy_specific_address(self, address):
        """Copy a specific address to clipboard"""
        self.root.clipboard_clear()
        self.root.clipboard_append(address)
        self.update_status("Address copied to clipboard", self.colors['success'])
        self.root.after(2000, lambda: self.update_status("Ready"))
    
    def google_search_specific_address(self, address):
        """Open Google search for a specific address"""
        search_url = f"https://www.google.com/search?q={urllib.parse.quote(address)}"
        webbrowser.open(search_url)
        self.update_status("Opening Google search...", self.colors['success'])
        self.root.after(2000, lambda: self.update_status("Ready"))
    
    def on_status_change(self, event=None):
        if not self.current_results:
            return
        
        new_status = self.status_var.get()
        self.current_results[self.current_result_idx]['status'] = new_status
        
        # Don't save to call history yet - wait until person is skipped
        
        self.update_status(f"Status updated to: {new_status}", self.colors['success'])
        self.update_data_in_excel()
        self.root.after(2000, lambda: self.update_status("Ready"))
    
    def on_notes_typing(self, event=None):
        """Handle notes text changes with debouncing"""
        if hasattr(self, '_notes_timer'):
            self.root.after_cancel(self._notes_timer)
        self._notes_timer = self.root.after(1000, self.on_notes_change)
    
    def on_notes_change(self, event=None):
        if not self.current_results:
            return
        
        new_notes = self.notes_text.get('1.0', tk.END).strip()
        self.current_results[self.current_result_idx]['notes'] = new_notes
        
        # Don't save to call history yet - wait until person is skipped
        
        self.update_data_in_excel()
        self.update_status("Notes saved", self.colors['success'])
        self.root.after(2000, lambda: self.update_status("Ready"))
    
    def update_data_in_excel(self):
        """Update status and notes in Excel"""
        try:
            if not os.path.exists(self.output_excel_file):
                return
            
            wb = load_workbook(self.output_excel_file)
            ws = wb.active
            
            result = self.current_results[self.current_result_idx]
            original_phone = result['original_phone']
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
                if str(row[1].value) == original_phone:
                    ws.cell(row=row_idx, column=9, value=result.get('status', ''))
                    ws.cell(row=row_idx, column=10, value=result.get('notes', ''))
                    break
            
            wb.save(self.output_excel_file)
            self.excel_cache.update(original_phone, result)
        except Exception as e:
            print(f"Error updating Excel: {e}")
    
    def show_cache_stats(self):
        """Show cache statistics dialog"""
        cache_manager = CacheManager()
        stats = cache_manager.get_statistics()
        
        info = f"Permanent Cache Statistics\n\n"
        info += f"Total Cached Entries: {stats['total_entries']}\n"
        info += f"Cache File Size: {stats['file_size_mb']} MB\n\n"
        info += f"Session Statistics:\n"
        info += f"  Cache Hits: {stats['cache_hits']}\n"
        info += f"  Cache Misses: {stats['cache_misses']}\n"
        info += f"  Hit Rate: {stats['hit_rate']}%\n\n"
        info += f"Cost Savings:\n"
        info += f"  API Calls Saved: {stats['api_calls_saved']}\n"
        info += f"  Estimated Savings: ${stats['estimated_savings_usd']}\n\n"
        
        if stats['oldest_entry']:
            info += f"Oldest Entry: {stats['oldest_entry'][:10]}\n"
        
        messagebox.showinfo("Cache Statistics", info)
    
    def clear_cache_confirm(self):
        """Confirm and clear cache"""
        cache_manager = CacheManager()
        stats = cache_manager.get_statistics()
        
        if stats['total_entries'] == 0:
            messagebox.showinfo("Cache Empty", "Cache is already empty.")
            return
        
        confirm = messagebox.askyesno(
            "Clear Cache",
            f"Clear all cached data?\n\n"
            f"This will remove {stats['total_entries']} cached entries.\n"
            f"Future lookups will require API calls and incur costs.\n\n"
            f"Are you sure?"
        )
        
        if confirm:
            entries_cleared, success = cache_manager.clear_cache()
            if success:
                messagebox.showinfo(
                    "Cache Cleared",
                    f"Successfully cleared {entries_cleared} cached entries.\n\n"
                    f"Future lookups will use the API."
                )
            else:
                messagebox.showerror("Error", "Failed to clear cache.")
    
    def html_escape(self, text):
        """Escape HTML special characters"""
        if not text:
            return ""
        return str(text).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;').replace('"', '&quot;').replace("'", '&#39;')
    
    def generate_visual_html(self, phone_data, address_data, is_from_cache):
        """Generate beautiful HTML for visual view"""
        try:
            # Extract people using the same extraction function
            people = []
            
            # Combine phone and address data
            combined_data = {
                'phone_lookup': phone_data,
                'address_lookup': address_data
            }
            
            # Use the extraction function
            people_from_data = self.extract_people_from_data(combined_data)
            
            # Convert to format expected by HTML
            for person_info in people_from_data:
                person = {
                    'name': self.html_escape(person_info['name']),
                    'phones': [{'number': self.html_escape(p), 'type': 'Mobile'} for p in person_info['phones']],
                    'addresses': [{'full': self.html_escape(a), 'type': 'Current'} for a in person_info['addresses']]
                }
                people.append(person)
        except Exception as e:
            print(f"Error extracting people: {e}")
            people = []
        
        total_phones = sum(len(p['phones']) for p in people)
        cache_indicator = "FROM CACHE" if is_from_cache else "FRESH FROM API"
        cache_color = "#27AE60" if is_from_cache else "#F39C12"
        
        html = f'''
<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            padding: 20px;
        }}
        .container {{
            max-width: 100%;
            background: white;
            border-radius: 12px;
            overflow: hidden;
        }}
        .header {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 20px;
            text-align: center;
        }}
        .header h1 {{ font-size: 24px; margin-bottom: 5px; }}
        .header p {{ opacity: 0.9; font-size: 12px; }}
        .cache-badge {{
            display: inline-block;
            background: {cache_color};
            color: white;
            padding: 5px 15px;
            border-radius: 20px;
            font-size: 11px;
            font-weight: 600;
            margin-top: 10px;
        }}
        .summary {{
            background: #e8eaf6;
            padding: 15px;
            margin: 15px;
            border-radius: 8px;
            border-left: 4px solid #667eea;
        }}
        .summary-stats {{
            display: flex;
            gap: 30px;
            font-size: 13px;
        }}
        .stat {{ font-weight: 600; }}
        .stat-value {{ color: #667eea; font-size: 16px; }}
        .results {{ padding: 15px; }}
        .person-card {{
            background: white;
            border: 2px solid #e0e0e0;
            border-radius: 8px;
            padding: 15px;
            margin-bottom: 15px;
        }}
        .person-card:hover {{
            border-color: #667eea;
            box-shadow: 0 4px 12px rgba(102, 126, 234, 0.2);
        }}
        .person-name {{
            font-size: 18px;
            font-weight: 700;
            color: #764ba2;
            margin-bottom: 12px;
        }}
        .address-section {{
            background: #e8eaf6;
            padding: 10px 12px;
            border-radius: 6px;
            margin-bottom: 10px;
            border-left: 3px solid #667eea;
        }}
        .address-label {{
            font-size: 11px;
            font-weight: 600;
            color: #666;
            margin-bottom: 4px;
            text-transform: uppercase;
        }}
        .address-text {{
            color: #333;
            font-size: 13px;
            line-height: 1.4;
        }}
        .phone-list {{
            display: flex;
            flex-direction: column;
            gap: 8px;
        }}
        .phone-item {{
            display: flex;
            justify-content: space-between;
            align-items: center;
            padding: 8px 12px;
            background: #f8f9fa;
            border-radius: 6px;
            border: 1px solid #e0e0e0;
        }}
        .phone-number {{
            font-family: 'Courier New', monospace;
            font-size: 14px;
            font-weight: 600;
            color: #333;
        }}
        .phone-type {{
            font-size: 11px;
            color: #666;
            background: #e8eaf6;
            padding: 3px 8px;
            border-radius: 4px;
            margin-left: 10px;
        }}
        .no-results {{
            text-align: center;
            padding: 40px;
            color: #666;
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>Passion Of Rugs Response Viewer</h1>
            <p>Extracted people, phone numbers, and addresses</p>
            <div class="cache-badge">{cache_indicator}</div>
        </div>
        
        <div class="summary">
            <div class="summary-stats">
                <div class="stat">
                    <div>People Found</div>
                    <div class="stat-value">{len(people)}</div>
                </div>
                <div class="stat">
                    <div>Total Phone Numbers</div>
                    <div class="stat-value">{total_phones}</div>
                </div>
            </div>
        </div>
        
        <div class="results">
'''
        
        if not people:
            html += '''
            <div class="no-results">
                <div style="font-size: 48px; margin-bottom: 10px;">🔍</div>
                <h3>No People Found</h3>
                <p>No names or phone numbers were found in the API response.</p>
            </div>
'''
        else:
            for person in people:
                html += f'''
            <div class="person-card">
                <div class="person-name">{person['name']}</div>
'''
                
                # Display addresses
                if person['addresses']:
                    for addr in person['addresses']:
                        html += f'''
                <div class="address-section">
                    <div class="address-label">{addr['type']} Address</div>
                    <div class="address-text">{addr['full']}</div>
                </div>
'''
                
                # Display phones
                if person['phones']:
                    html += '<div class="phone-list">'
                    for phone in person['phones']:
                        html += f'''
                <div class="phone-item">
                    <div>
                        <span class="phone-number">{phone['number']}</span>
                        <span class="phone-type">{phone['type']}</span>
                    </div>
                </div>
'''
                    html += '</div>'
                else:
                    html += '<p style="color: #666; font-style: italic; font-size: 13px;">No phone numbers available</p>'
                
                html += '</div>'
        
        html += '''
        </div>
    </div>
</body>
</html>
'''
        
        return html
    
    def show_full_response(self):
        """Show full API response in a popup window with HTML viewer"""
        if not self.current_results or self.current_result_idx >= len(self.current_results):
            messagebox.showwarning("No Data", "No result data available to display.")
            return
        
        # Get current person's original data
        if self.current_person_idx >= len(self.original_data):
            return
        
        original_person = self.original_data[self.current_person_idx]
        original_phone = self.lead_processor.clean_phone(original_person.get('phone', ''))
        
        # Get cached API responses
        phone_data = {}
        address_data = {}
        is_from_cache = False
        
        if self.lead_processor.cache_manager:
            cached = self.lead_processor.cache_manager.get_cached_lookup(original_phone)
            if cached:
                phone_data, address_data = cached
                is_from_cache = True
        
        # Create popup window
        response_win = tk.Toplevel(self.root)
        response_win.title("Full API Response")
        response_win.geometry("1000x700")
        response_win.configure(bg=self.colors['bg'])
        
        # Title with cache indicator
        title_text = "Full API Response Viewer"
        if is_from_cache:
            title_text += " (FROM CACHE)"
        else:
            title_text += " (FROM API)"
        
        title_label = tk.Label(
            response_win,
            text=title_text,
            font=('Arial', 16, 'bold'),
            bg=self.colors['bg'],
            fg=self.colors['success'] if is_from_cache else self.colors['warning']
        )
        title_label.pack(pady=10)
        
        # Tabs for different views
        notebook = ttk.Notebook(response_win)
        notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Tab 1: Formatted View
        formatted_frame = tk.Frame(notebook, bg='white')
        notebook.add(formatted_frame, text="Formatted View")
        
        formatted_text = scrolledtext.ScrolledText(
            formatted_frame,
            font=('Courier New', 10),
            wrap=tk.WORD,
            bg='white'
        )
        formatted_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Format the data nicely
        formatted_content = "=" * 80 + "\n"
        formatted_content += "PHONE LOOKUP RESPONSE\n"
        formatted_content += "=" * 80 + "\n\n"
        formatted_content += json.dumps(phone_data, indent=2)
        formatted_content += "\n\n" + "=" * 80 + "\n"
        formatted_content += "ADDRESS LOOKUP RESPONSE\n"
        formatted_content += "=" * 80 + "\n\n"
        formatted_content += json.dumps(address_data, indent=2)
        
        formatted_text.insert('1.0', formatted_content)
        formatted_text.config(state='disabled')
        
        # Tab 2: Raw JSON
        raw_frame = tk.Frame(notebook, bg='white')
        notebook.add(raw_frame, text="Raw JSON")
        
        raw_text = scrolledtext.ScrolledText(
            raw_frame,
            font=('Courier New', 10),
            wrap=tk.WORD,
            bg='white'
        )
        raw_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        raw_content = {
            'phone_lookup': phone_data,
            'address_lookup': address_data
        }
        raw_text.insert('1.0', json.dumps(raw_content, indent=2))
        raw_text.config(state='disabled')
        
        # Tab 3: Extracted Data
        extracted_frame = tk.Frame(notebook, bg='white')
        notebook.add(extracted_frame, text="Extracted Data")
        
        extracted_text = scrolledtext.ScrolledText(
            extracted_frame,
            font=('Courier New', 10),
            wrap=tk.WORD,
            bg='white'
        )
        extracted_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Extract and display people found
        extracted_content = "=" * 80 + "\n"
        extracted_content += "EXTRACTED PEOPLE AND CONTACT INFO\n"
        if is_from_cache:
            extracted_content += "(Loaded from Cache)\n"
        else:
            extracted_content += "(Fresh from API)\n"
        extracted_content += "=" * 80 + "\n\n"
        
        for idx, result in enumerate(self.current_results, 1):
            extracted_content += f"Person {idx}:\n"
            extracted_content += f"  Name: {result.get('new_name', 'N/A')}\n"
            extracted_content += f"  Age: {result.get('age', 'N/A')}\n"
            
            # Get phones
            phones = result.get('new_phones', [])
            if phones:
                extracted_content += f"  Phones: {', '.join(phones)}\n"
            else:
                extracted_content += f"  Phones: None\n"
            
            # Get addresses
            addresses = result.get('new_addresses', [])
            if addresses:
                extracted_content += f"  Addresses:\n"
                for addr in addresses:
                    extracted_content += f"    - {addr}\n"
            else:
                extracted_content += f"  Addresses: None\n"
            
            extracted_content += "\n"
        
        extracted_text.insert('1.0', extracted_content)
        extracted_text.config(state='disabled')
        
        # Tab 4: Visual HTML View (if tkinterweb is available)
        if HAS_TKINTERWEB:
            try:
                # Generate beautiful HTML first
                html_content = self.generate_visual_html(phone_data, address_data, is_from_cache)
                
                # Create HTML frame
                html_frame = tkinterweb.HtmlFrame(notebook, messages_enabled=False)
                notebook.add(html_frame, text="Visual View")
                
                # Load HTML in a safe way
                try:
                    html_frame.load_html(html_content)
                except Exception as load_error:
                    print(f"Error loading HTML: {load_error}")
                    # Show error in the frame
                    error_html = f'''
                    <html><body style="font-family: Arial; padding: 20px;">
                    <h3 style="color: #c62828;">Error Loading Visual View</h3>
                    <p>There was an error rendering the HTML view.</p>
                    <p style="color: #666; font-size: 12px;">Error: {str(load_error)}</p>
                    <p>Please use the other tabs to view the data.</p>
                    </body></html>
                    '''
                    html_frame.load_html(error_html)
            except Exception as e:
                print(f"Error creating HTML view: {e}")
                # Create fallback frame with error message
                error_frame = tk.Frame(notebook, bg='white')
                notebook.add(error_frame, text="Visual View")
                
                error_msg = tk.Label(
                    error_frame,
                    text=f"Visual HTML View Error\n\n"
                         f"An error occurred while creating the visual view:\n\n"
                         f"{str(e)}\n\n"
                         f"Please use the other tabs to view the data.",
                    font=('Arial', 11),
                    bg='white',
                    fg='#c62828',
                    justify='center'
                )
                error_msg.pack(expand=True, pady=50)
        else:
            # Show message about installing tkinterweb
            install_frame = tk.Frame(notebook, bg='white')
            notebook.add(install_frame, text="Visual View")
            
            install_msg = tk.Label(
                install_frame,
                text="Visual HTML View Not Available\n\n"
                     "To enable the beautiful visual view, install tkinterweb:\n\n"
                     "pip install tkinterweb\n\n"
                     "Then restart the application.",
                font=('Arial', 12),
                bg='white',
                fg='#666',
                justify='center'
            )
            install_msg.pack(expand=True, pady=50)
        
        # Close button
        tk.Button(
            response_win,
            text="Close",
            command=response_win.destroy,
            font=('Arial', 11, 'bold'),
            bg=self.colors['secondary'],
            fg='white',
            padx=30,
            pady=10,
            cursor='hand2'
        ).pack(pady=10)
    
    def manual_phone_lookup(self):
        """Manually trigger phone lookup - always calls API, accumulates with existing data"""
        if self.current_person_idx >= len(self.original_data):
            return
        
        original_person = self.original_data[self.current_person_idx]
        original_phone = self.lead_processor.clean_phone(original_person.get('phone', ''))
        
        if not original_phone:
            messagebox.showwarning("No Phone", "No phone number available for lookup.")
            return
        
        self.update_status("Performing manual phone lookup from API...", self.colors['warning'])
        
        def do_lookup():
            try:
                # Always call API directly (bypass cache)
                phone_data = self.lead_processor.phone_lookup(original_phone)
                
                # Get existing address data from cache to preserve it
                address_data = {}
                if self.lead_processor.cache_manager:
                    cached = self.lead_processor.cache_manager.get_cached_lookup(original_phone)
                    if cached:
                        address_data = cached[1] or {}
                
                # Process and accumulate results
                self.root.after(0, lambda: self.accumulate_phone_results(phone_data, original_person))
                
                # Update cache with new phone data + existing address data
                if self.lead_processor.cache_manager:
                    self.lead_processor.cache_manager.store_lookup(original_phone, phone_data, address_data)
                
                self.root.after(0, lambda: self.update_status("Manual phone lookup complete - data added", self.colors['success']))
            except Exception as e:
                error_msg = f"Phone lookup error: {str(e)}"
                self.root.after(0, lambda: self.update_status(error_msg, self.colors['danger']))
        
        threading.Thread(target=do_lookup, daemon=True).start()
    
    def manual_address_lookup(self):
        """Manually trigger address lookup - always calls API, accumulates with existing data"""
        if self.current_person_idx >= len(self.original_data):
            return
        
        original_person = self.original_data[self.current_person_idx]
        original_phone = self.lead_processor.clean_phone(original_person.get('phone', ''))
        
        # Parse address
        street = original_person.get('address', '')
        city = original_person.get('city', '')
        state = original_person.get('state', '')
        zip_code = original_person.get('zip', '')
        
        if street and (not city or not state or not zip_code):
            parsed = self.lead_processor.parse_address(street)
            street = parsed['street'] or street
            city = parsed['city'] or city
            state = parsed['state'] or state
            zip_code = parsed['zip'] or zip_code
        
        if not any([street, city, state, zip_code]):
            messagebox.showwarning("No Address", "No address information available for lookup.")
            return
        
        self.update_status("Performing manual address lookup from API...", self.colors['warning'])
        
        def do_lookup():
            try:
                # Always call API directly (bypass cache)
                address_data = self.lead_processor.address_lookup(street, city, state, zip_code)
                
                # Get existing phone data from cache to preserve it
                phone_data = {}
                if self.lead_processor.cache_manager:
                    cached = self.lead_processor.cache_manager.get_cached_lookup(original_phone)
                    if cached:
                        phone_data = cached[0] or {}
                
                # Process and accumulate results
                self.root.after(0, lambda: self.accumulate_address_results(address_data, original_person))
                
                # Update cache with existing phone data + new address data
                if self.lead_processor.cache_manager:
                    self.lead_processor.cache_manager.store_lookup(original_phone, phone_data, address_data)
                
                self.root.after(0, lambda: self.update_status("Manual address lookup complete - data added", self.colors['success']))
            except Exception as e:
                error_msg = f"Address lookup error: {str(e)}"
                self.root.after(0, lambda: self.update_status(error_msg, self.colors['danger']))
        
        threading.Thread(target=do_lookup, daemon=True).start()
    
    def save_note_with_status(self):
        """Save note with current status to call history"""
        if not self.current_results or self.current_result_idx >= len(self.current_results):
            return
        
        current_result = self.current_results[self.current_result_idx]
        phone = current_result.get('phones', [''])[self.current_phone_idx] if current_result.get('phones') else ''
        name = current_result.get('name', '')
        status = self.status_var.get()
        notes = self.notes_text.get('1.0', tk.END).strip()
        
        if not status:
            messagebox.showwarning("No Status", "Please select a status before saving note.")
            return
        
        # Update current result
        current_result['status'] = status
        current_result['notes'] = notes
        
        # Add to call history
        self.add_to_call_history(phone, name, status, notes)
        
        # Save to Excel
        self.update_data_in_excel()
        
        self.update_status(f"Note saved with status: {status}", self.colors['success'])
    
    def flash_call_button(self):
        """Flash the call button after clicking"""
        original_bg = self.call_btn.cget('bg')
        original_text = self.call_btn.cget('text')
        
        def flash(count=0):
            if count < 6:  # Flash 3 times (6 state changes)
                if count % 2 == 0:
                    self.call_btn.config(bg='#FFD700', text='📞 Calling...')
                else:
                    self.call_btn.config(bg=original_bg, text=original_text)
                self.root.after(200, lambda: flash(count + 1))
            else:
                self.call_btn.config(bg=original_bg, text=original_text)
        
        flash()
    
    def show_call_history_window(self):
        """Show call history in a popup window"""
        history_win = tk.Toplevel(self.root)
        history_win.title("Call History")
        history_win.geometry("900x600")
        history_win.configure(bg=self.colors['bg'])
        
        # Title
        tk.Label(
            history_win,
            text="Call History",
            font=('Arial', 14, 'bold'),
            bg=self.colors['bg'],
            fg=self.colors['primary']
        ).pack(pady=10)
        
        # Info bar
        info_frame = tk.Frame(history_win, bg=self.colors['light'], relief=tk.SUNKEN, bd=1)
        info_frame.pack(fill=tk.X, padx=10, pady=(0, 5))
        
        tk.Label(
            info_frame,
            text=f"Total Calls: {len(self.call_history)}",
            font=('Arial', 9, 'bold'),
            bg=self.colors['light'],
            fg=self.colors['primary']
        ).pack(side=tk.LEFT, padx=10, pady=5)
        
        # Create Treeview for call history
        tree_frame = tk.Frame(history_win, bg=self.colors['bg'])
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        # Scrollbars
        tree_scroll_y = ttk.Scrollbar(tree_frame, orient="vertical")
        tree_scroll_x = ttk.Scrollbar(tree_frame, orient="horizontal")
        
        # Treeview
        tree = ttk.Treeview(
            tree_frame,
            columns=('Timestamp', 'Phone', 'Name', 'Status', 'Notes'),
            show='headings',
            yscrollcommand=tree_scroll_y.set,
            xscrollcommand=tree_scroll_x.set
        )
        
        tree_scroll_y.config(command=tree.yview)
        tree_scroll_x.config(command=tree.xview)
        
        # Define columns
        tree.heading('Timestamp', text='Date & Time')
        tree.heading('Phone', text='Phone Number')
        tree.heading('Name', text='Name')
        tree.heading('Status', text='Status')
        tree.heading('Notes', text='Notes')
        
        tree.column('Timestamp', width=150, anchor='w')
        tree.column('Phone', width=120, anchor='w')
        tree.column('Name', width=150, anchor='w')
        tree.column('Status', width=100, anchor='w')
        tree.column('Notes', width=300, anchor='w')
        
        # Pack scrollbars and tree
        tree_scroll_y.pack(side=tk.RIGHT, fill=tk.Y)
        tree_scroll_x.pack(side=tk.BOTTOM, fill=tk.X)
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # Enable mouse wheel scrolling for call history
        def on_tree_mousewheel(event):
            tree.yview_scroll(int(-1*(event.delta/120)), "units")
        tree.bind("<Enter>", lambda e: tree.bind_all("<MouseWheel>", on_tree_mousewheel))
        tree.bind("<Leave>", lambda e: tree.unbind_all("<MouseWheel>"))
        
        # Populate tree with call history (most recent first)
        for entry in reversed(self.call_history):
            timestamp = entry.get('timestamp', '')
            # Format timestamp for display
            try:
                dt = datetime.datetime.fromisoformat(timestamp)
                formatted_time = dt.strftime('%Y-%m-%d %H:%M:%S')
            except:
                formatted_time = timestamp
            
            tree.insert('', 'end', values=(
                formatted_time,
                entry.get('phone', ''),
                entry.get('name', ''),
                entry.get('status', ''),
                entry.get('notes', '')
            ))
        
        # Buttons frame
        button_frame = tk.Frame(history_win, bg=self.colors['bg'])
        button_frame.pack(pady=10)
        
        # Export button
        tk.Button(
            button_frame,
            text="📥 Export to Excel",
            command=lambda: self.export_call_history(),
            font=('Arial', 9),
            bg=self.colors['secondary'],
            fg='white',
            padx=15,
            pady=6,
            cursor='hand2'
        ).pack(side=tk.LEFT, padx=5)
        
        # Clear button
        tk.Button(
            button_frame,
            text="🗑️ Clear History",
            command=lambda: self.clear_call_history(history_win),
            font=('Arial', 9),
            bg=self.colors['danger'],
            fg='white',
            padx=15,
            pady=6,
            cursor='hand2'
        ).pack(side=tk.LEFT, padx=5)
        
        # Close button
        tk.Button(
            button_frame,
            text="Close",
            command=history_win.destroy,
            font=('Arial', 9),
            bg=self.colors['light'],
            padx=15,
            pady=6,
            cursor='hand2'
        ).pack(side=tk.LEFT, padx=5)
    
    def export_call_history(self):
        """Export call history to Excel file"""
        if not self.call_history:
            messagebox.showinfo("No Data", "Call history is empty.")
            return
        
        try:
            # Create DataFrame
            df = pd.DataFrame(self.call_history)
            
            # Format timestamp
            if 'timestamp' in df.columns:
                df['timestamp'] = pd.to_datetime(df['timestamp']).dt.strftime('%Y-%m-%d %H:%M:%S')
            
            # Save to Excel
            output_file = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile="call_history.xlsx"
            )
            
            if output_file:
                df.to_excel(output_file, index=False)
                messagebox.showinfo("Success", f"Call history exported to:\n{output_file}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export call history:\n{str(e)}")
    
    def clear_call_history(self, window=None):
        """Clear all call history after confirmation"""
        if not self.call_history:
            messagebox.showinfo("No Data", "Call history is already empty.")
            return
        
        response = messagebox.askyesno(
            "Confirm Clear",
            f"Are you sure you want to clear all {len(self.call_history)} call history entries?\n\nThis action cannot be undone."
        )
        
        if response:
            self.call_history = []
            self.save_call_history()
            messagebox.showinfo("Success", "Call history cleared.")
            if window:
                window.destroy()


    def process_ai_analysis(self, original_data, phone_response, address_response):
        """Process AI analysis for current person"""
        if not self.ai_assistant or not self.settings.get('ai_enabled'):
            return None
        
        if not self.settings.get('ai_person_filtering'):
            return None
        
        # Extract original data
        original_name = original_data.get('name', '')
        original_phone = original_data.get('phone', '')
        original_address = original_data.get('address', '')
        
        # Run AI filtering
        try:
            self.update_status("AI analyzing results...", self.colors['secondary'])
            ai_results = self.ai_assistant.filter_and_rank_contacts(
                original_name, original_phone, original_address,
                phone_response or {}, address_response or {}
            )
            
            if ai_results:
                self.update_status("AI analysis complete", self.colors['success'])
                return ai_results
            else:
                self.update_status("AI analysis failed - showing raw data", self.colors['warning'])
                return None
        except Exception as e:
            logger.error(f"AI analysis error: {e}")
            self.update_status("AI analysis error", self.colors['danger'])
            return None
    
    def show_ai_tab(self, parent_frame):
        """Create and populate AI Overview tab - validates AI results match current person"""
        # Validate we have a valid person index
        if self.current_person_idx < 0 or self.current_person_idx >= len(self.original_data):
            logger.warning(f"Invalid person index in show_ai_tab: {self.current_person_idx}")
            return
        
        # Clear existing content
        for widget in parent_frame.winfo_children():
            widget.destroy()
        
        # Create scrollable frame
        canvas = tk.Canvas(parent_frame, bg='white', highlightthickness=0)
        scrollbar = ttk.Scrollbar(parent_frame, orient="vertical", command=canvas.yview)
        scroll_frame = tk.Frame(canvas, bg='white')
        
        canvas.create_window((0, 0), window=scroll_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Enable mouse wheel scrolling for AI tab
        def on_ai_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        canvas.bind("<Enter>", lambda e: canvas.bind_all("<MouseWheel>", on_ai_mousewheel))
        canvas.bind("<Leave>", lambda e: canvas.unbind_all("<MouseWheel>"))
        
        # Check if AI is enabled
        if not self.ai_assistant or not self.settings.get('ai_enabled'):
            tk.Label(
                scroll_frame,
                text="AI Features Disabled",
                font=('Arial', 14, 'bold'),
                bg='white',
                fg='#999'
            ).pack(pady=50)
            
            tk.Label(
                scroll_frame,
                text="Configure OpenAI API key in setup to enable AI features",
                font=('Arial', 10),
                bg='white',
                fg='#666'
            ).pack(pady=10)
            return
        
        # Validate AI results match current person
        current_person = self.original_data[self.current_person_idx]
        current_phone = self.lead_processor.clean_phone(current_person.get('phone', ''))
        
        # Check if AI results exist and validate they're for the current person
        if not self.ai_results:
            tk.Label(
                scroll_frame,
                text=dy=50)
            
            tk.Label(
                scroll_frame,
                text="AI analysis will appear here after API lookups complete",
                font=('Arial', 10),
                bg='white',
                fg='#666'
            ).pack(pady=10)
            return
        
        # Display AI results
        padding = {'padx': 15, 'pady': 8}
        
        # Original Customer Data section
        orig_section = tk.LabelFrame(
            scroll_frame,
            text="ORIGINAL CUSTOMER DATA",
            font=('Arial', 10, 'bold'),
            bg='white',
            fg=self.colors['primary']
        )
        orig_section.pack(fill=tk.X, **padding)
        
        if self.current_person_idx < len(self.original_data):
            person = self.original_data[self.current_person_idx]
            
            info_frame = tk.Frame(orig_section, bg='white')
            info_frame.pack(fill=tk.X, padx=10, pady=8)
            
            tk.Label(info_frame, text=f"Name: {person.get('name', 'N/A')}", 
                    font=('Arial', 9), bg='white').pack(anchor='w', pady=2)
            tk.Label(info_frame, text=f"Phone: {person.get('phone', 'N/A')}", 
                    font=('Arial', 9), bg='white').pack(anchor='w', pady=2)
            tk.Label(info_frame, text=f"Address: {person.get('address', 'N/A')}", 
                    font=('Arial', 9), bg='white').pack(anchor='w', pady=2)
            
            # Show AI correction if any
            if self.ai_correction_log:
                for log_entry in self.ai_correction_log:
                    if log_entry.get('person_idx') == self.current_person_idx:
                        tk.Label(
                            info_frame,
                            text=f"✓ AI Corrected: {log_entry.get('corrected_address', '')}",
                            font=('Arial', 9, 'italic'),
                            bg='white',
                            fg=self.colors['success']
                        ).pack(anchor='w', pady=2)
        
        # AI Horizontal Ranking section
        ranking_section = tk.LabelFrame(
            scroll_frame,
            text="AI HORIZONTAL RANKING - CALLING STRATEGY",
            font=('Arial', 10, 'bold'),
            bg='white',
            fg=self.colors['primary']
        )
        ranking_section.pack(fill=tk.X, **padding)
        
        ranking_frame = tk.Frame(ranking_section, bg='white')
        ranking_frame.pack(fill=tk.X, padx=10, pady=8)
        
        # Display horizontal ranking
        horizontal_ranking = self.ai_results.get('horizontal_ranking', [])
        if horizontal_ranking:
            # Create horizontal layout for persons
            for person in horizontal_ranking:
                self._display_ranked_person(ranking_frame, person)
        
        # Calling Strategy section
        strategy = self.ai_results.get('calling_strategy', {})
        if strategy:
            strategy_section = tk.LabelFrame(
                scroll_frame,
                text="AI CALLING STRATEGY & RECOMMENDATIONS",
                font=('Arial', 10, 'bold'),
                bg='white',
                fg=self.colors['primary']
            )
            strategy_section.pack(fill=tk.X, **padding)
            
            strategy_frame = tk.Frame(strategy_section, bg='white')
            strategy_frame.pack(fill=tk.X, padx=10, pady=8)
            
            # Primary recommendation
            primary_rec = strategy.get('primary_recommendation', '')
            if primary_rec:
                tk.Label(
                    strategy_frame,
                    text=f"🎯 {primary_rec}",
                    font=('Arial', 10, 'bold'),
                    bg='white',
                    fg=self.colors['success']
                ).pack(anchor='w', pady=3)
            
            # Backup plan
            backup_plan = strategy.get('backup_plan', '')
            if backup_plan:
                tk.Label(
                    strategy_frame,
                    text=f"🔄 Backup Plan: {backup_plan}",
                    font=('Arial', 9),
                    bg='white',
                    wraplength=600,
                    justify='left'
                ).pack(anchor='w', pady=3)
            
            # Best time to call
            best_time = strategy.get('best_time_to_call', '')
            if best_time:
                tk.Label(
                    strategy_frame,
                    text=f"⏰ Best Time to Call: {best_time}",
                    font=('Arial', 9),
                    bg='white'
                ).pack(anchor='w', pady=3)
            
            # Acceptance probability
            prob = strategy.get('acceptance_probability', 0)
            if prob:
                tk.Label(
                    strategy_frame,
                    text=f"📊 Success Probability: {prob}% ({'High' if prob >= 80 else 'Medium' if prob >= 50 else 'Low'})",
                    font=('Arial', 9, 'bold'),
                    bg='white',
                    fg=self.colors['success'] if prob >= 80 else self.colors['warning']
                ).pack(anchor='w', pady=3)
            
            # Overall reasoning
            reasoning = strategy.get('overall_reasoning', '')
            if reasoning:
                tk.Label(
                    strategy_frame,
                    text=f"💡 Strategy: {reasoning}",
                    font=('Arial', 9),
                    bg='white',
                    fg='#666',
                    wraplength=600,
                    justify='left'
                ).pack(anchor='w', pady=3)
        
        # Status and Notes section (same as Standard View)
        self._add_ai_status_notes_section(scroll_frame)
        
        # Update scroll region
        scroll_frame.update_idletasks()
        canvas.config(scrollregion=canvas.bbox("all"))
    
    def _display_ranked_person(self, parent, person):
        """Display a single ranked person with all their phone numbers in horizontal layout"""
        rank = person.get('rank', 0)
        person_name = person.get('person_name', 'Unknown')
        relationship = person.get('relationship', 'Unknown')
        confidence_score = person.get('confidence_score', 0)
        confidence_level = person.get('confidence_level', 'Low')
        reasoning = person.get('reasoning', '')
        current_address = person.get('current_address', '')
        all_phone_numbers = person.get('all_phone_numbers', [])
        recommended_first_call = person.get('recommended_first_call', '')
        
        # Main person frame with ranking
        person_frame = tk.Frame(parent, bg='#f0f8ff', relief=tk.RAISED, bd=2)
        person_frame.pack(fill=tk.X, pady=8, padx=5)
        
        # Header with rank and person info
        header_frame = tk.Frame(person_frame, bg='#f0f8ff')
        header_frame.pack(fill=tk.X, padx=10, pady=8)
        
        # Rank badge
        rank_color = self.colors['success'] if rank == 1 else self.colors['warning'] if rank == 2 else self.colors['secondary']
        tk.Label(
            header_frame,
            text=f"#{rank}",
            font=('Arial', 12, 'bold'),
            bg=rank_color,
            fg='white',
            width=3,
            relief=tk.RAISED,
            bd=1
        ).pack(side=tk.LEFT, padx=(0, 10))
        
        # Person details
        details_frame = tk.Frame(header_frame, bg='#f0f8ff')
        details_frame.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        # Name and relationship
        name_text = f"{person_name}"
        if relationship != 'Unknown':
            name_text += f" ({relationship})"
        tk.Label(
            details_frame,
            text=name_text,
            font=('Arial', 11, 'bold'),
            bg='#f0f8ff'
        ).pack(anchor='w')
        
        # Confidence score
        confidence_color = self.colors['success'] if confidence_level == 'High' else self.colors['warning'] if confidence_level == 'Medium' else self.colors['danger']
        tk.Label(
            details_frame,
            text=f"Confidence: {confidence_score}% ({confidence_level})",
            font=('Arial', 9, 'bold'),
            bg='#f0f8ff',
            fg=confidence_color
        ).pack(anchor='w')
        
        # Reasoning
        if reasoning:
            tk.Label(
                details_frame,
                text=f"└─ {reasoning}",
                font=('Arial', 8),
                bg='#f0f8ff',
                fg='#666',
                wraplength=400,
                justify='left'
            ).pack(anchor='w', padx=10)
        
        # Current address
        if current_address:
            tk.Label(
                details_frame,
                text=f"📍 {current_address}",
                font=('Arial', 8),
                bg='#f0f8ff',
                fg='#666'
            ).pack(anchor='w', pady=(2, 0))
        
        # Phone numbers section
        if all_phone_numbers:
            phones_frame = tk.Frame(person_frame, bg='#f0f8ff')
            phones_frame.pack(fill=tk.X, padx=10, pady=(0, 8))
            
            tk.Label(
                phones_frame,
                text="📞 All Phone Numbers (ranked by call priority):",
                font=('Arial', 9, 'bold'),
                bg='#f0f8ff'
            ).pack(anchor='w', pady=(5, 3))
            
            # Display each phone number
            for phone_info in all_phone_numbers:
                self._display_phone_option(phones_frame, phone_info, recommended_first_call)
    
    def _display_phone_option(self, parent, phone_info, recommended_first_call):
        """Display a single phone option with call and copy buttons"""
        phone = phone_info.get('phone', '')
        phone_type = phone_info.get('phone_type', '')
        carrier = phone_info.get('carrier', '')
        call_priority = phone_info.get('call_priority', 0)
        call_reasoning = phone_info.get('call_reasoning', '')
        
        # Phone frame
        phone_frame = tk.Frame(parent, bg='#ffffff', relief=tk.RAISED, bd=1)
        phone_frame.pack(fill=tk.X, pady=2, padx=10)
        
        # Left side - phone info
        info_frame = tk.Frame(phone_frame, bg='#ffffff')
        info_frame.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=8, pady=5)
        
        # Priority indicator
        priority_color = self.colors['success'] if call_priority == 1 else self.colors['warning'] if call_priority == 2 else self.colors['secondary']
        is_recommended = phone == recommended_first_call
        
        priority_text = f"Priority {call_priority}"
        if is_recommended:
            priority_text += " ⭐ RECOMMENDED"
        
        tk.Label(
            info_frame,
            text=priority_text,
            font=('Arial', 8, 'bold'),
            bg='#ffffff',
            fg=priority_color
        ).pack(anchor='w')
        
        # Phone number with type
        phone_text = f"📱 {phone}"
        if phone_type:
            phone_text += f" ({phone_type}"
            if carrier:
                phone_text += f" - {carrier}"
            phone_text += ")"
        
        tk.Label(
            info_frame,
            text=phone_text,
            font=('Arial', 9, 'bold' if is_recommended else 'normal'),
            bg='#ffffff'
        ).pack(anchor='w')
        
        # Call reasoning
        if call_reasoning:
            tk.Label(
                info_frame,
                text=f"└─ {call_reasoning}",
                font=('Arial', 8),
                bg='#ffffff',
                fg='#666'
            ).pack(anchor='w', padx=10)
        
        # Right side - action buttons
        buttons_frame = tk.Frame(phone_frame, bg='#ffffff')
        buttons_frame.pack(side=tk.RIGHT, padx=8, pady=5)
        
        # Copy button
        tk.Button(
            buttons_frame,
            text="Copy",
            command=lambda: self.copy_to_clipboard(phone),
            font=('Arial', 8),
            bg=self.colors['secondary'],
            fg='white',
            padx=8,
            pady=2,
            cursor='hand2'
        ).pack(side=tk.LEFT, padx=2)
        
        # Call button
        call_bg = self.colors['success'] if is_recommended else self.colors['primary']
        
        # Get person name from current results for call tracking
        person_name = 'AI Contact'
        if self.current_results and self.current_result_idx < len(self.current_results):
            person_name = self.current_results[self.current_result_idx].get('new_name', 'AI Contact')
        
        tk.Button(
            buttons_frame,
            text="Call via CloudTalk",
            command=lambda p=phone, n=person_name: self.make_call(p, n),
            font=('Arial', 8, 'bold' if is_recommended else 'normal'),
            bg=call_bg,
            fg='white',
            padx=8,
            pady=2,
            cursor='hand2'
        ).pack(side=tk.LEFT, padx=2)
    
    def _add_ai_status_notes_section(self, parent):
        """Add status and notes section to AI tab (same functionality as Standard View)"""
        # Status and Notes section
        status_notes_section = tk.LabelFrame(
            parent,
            text="CALL STATUS & NOTES",
            font=('Arial', 10, 'bold'),
            bg='white',
            fg=self.colors['primary']
        )
        status_notes_section.pack(fill=tk.X, padx=15, pady=8)
        
        status_notes_frame = tk.Frame(status_notes_section, bg='white')
        status_notes_frame.pack(fill=tk.X, padx=10, pady=8)
        
        # Status
        tk.Label(status_notes_frame, text="Status:", font=('Arial', 9, 'bold'), 
                bg='white').pack(side=tk.LEFT, padx=(0, 5))
        
        # Use the same status_var and status_dropdown as Standard View
        if not hasattr(self, 'status_var'):
            self.status_var = tk.StringVar()
        
        # Create AI tab status dropdown (linked to the same variable)
        self.ai_status_dropdown = ttk.Combobox(
            status_notes_frame,
            textvariable=self.status_var,
            values=['', 'No answer', 'Not interested', 'Wrong Number', 'Call back', 'Appointment', 'DNC'],
            font=('Arial', 10),
            width=18,
            state='readonly'
        )
        self.ai_status_dropdown.pack(side=tk.LEFT, padx=(0, 15))
        self.ai_status_dropdown.bind('<<ComboboxSelected>>', self.on_status_change)
        
        # Notes
        tk.Label(status_notes_frame, text="Notes:", font=('Arial', 9, 'bold'), 
                bg='white').pack(side=tk.LEFT, padx=(0, 5))
        
        # Use the same notes_text as Standard View
        if not hasattr(self, 'notes_text'):
            self.notes_text = tk.Text(
                status_notes_frame,
                font=('Arial', 9),
                width=30,
                height=2,
                wrap=tk.WORD,
                relief=tk.SOLID,
                borderwidth=1
            )
        
        # Create AI tab notes text (linked to the same functionality)
        self.ai_notes_text = tk.Text(
            status_notes_frame,
            font=('Arial', 9),
            width=30,
            height=2,
            wrap=tk.WORD,
            relief=tk.SOLID,
            borderwidth=1
        )
        self.ai_notes_text.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        
        # Sync AI notes with main notes
        def sync_ai_notes_to_main(event=None):
            if hasattr(self, 'notes_text'):
                content = self.ai_notes_text.get('1.0', tk.END)
                self.notes_text.delete('1.0', tk.END)
                self.notes_text.insert('1.0', content)
                self.on_notes_change(event)
        
        def sync_main_notes_to_ai():
            if hasattr(self, 'notes_text'):
                content = self.notes_text.get('1.0', tk.END)
                self.ai_notes_text.delete('1.0', tk.END)
                self.ai_notes_text.insert('1.0', content)
        
        self.ai_notes_text.bind('<FocusOut>', sync_ai_notes_to_main)
        self.ai_notes_text.bind('<KeyRelease>', lambda e: self.on_notes_typing(e) if hasattr(self, 'on_notes_typing') else None)
        
        # Store sync function for later use
        self.sync_main_notes_to_ai = sync_main_notes_to_ai
        
        # Save Note button
        tk.Button(
            status_notes_frame,
            text="💾",
            command=self.save_note_with_status,
            font=('Arial', 10),
            bg=self.colors['success'],
            fg='white',
            padx=8,
            pady=2,
            cursor='hand2'
        ).pack(side=tk.LEFT, padx=5)
        
        # Sync current data to AI tab
        if hasattr(self, 'notes_text'):
            sync_main_notes_to_ai()
    
    def _display_ai_contact(self, parent, contact, is_primary=True):
        """Display a single AI contact entry"""
        contact_frame = tk.Frame(parent, bg='#f8f9fa', relief=tk.RAISED, bd=1)
        contact_frame.pack(fill=tk.X, pady=5)
        
        inner = tk.Frame(contact_frame, bg='#f8f9fa')
        inner.pack(fill=tk.X, padx=10, pady=8)
        
        # Name and confidence
        name = contact.get('name', 'Unknown')
        confidence = contact.get('confidence_score', 0)
        confidence_level = contact.get('confidence_level', 'Low')
        
        color = self.colors['success'] if confidence >= 80 else self.colors['warning'] if confidence >= 50 else self.colors['danger']
        
        header_text = f"✓ {name} - Confidence: {confidence}% ({confidence_level})"
        if not is_primary:
            relationship = contact.get('relationship', '')
            header_text = f"{name} ({relationship}) - Confidence: {confidence}% ({confidence_level})"
        
        tk.Label(
            inner,
            text=header_text,
            font=('Arial', 9, 'bold'),
            bg='#f8f9fa',
            fg=color
        ).pack(anchor='w')
        
        # Phone
        phone = contact.get('phone', '')
        phone_type = contact.get('phone_type', '')
        carrier = contact.get('carrier', '')
        
        phone_text = f"📱 {phone}"
        if phone_type:
            phone_text += f" ({phone_type}"
            if carrier:
                phone_text += f" - {carrier}"
            phone_text += ")"
        
        phone_frame = tk.Frame(inner, bg='#f8f9fa')
        phone_frame.pack(anchor='w', pady=3)
        
        tk.Label(
            phone_frame,
            text=phone_text,
            font=('Arial', 9),
            bg='#f8f9fa'
        ).pack(side=tk.LEFT)
        
        # Copy and Call buttons
        if phone:
            tk.Button(
                phone_frame,
                text="📋 Copy",
                command=lambda p=phone: self._copy_to_clipboard(p),
                font=('Arial', 7),
                bg=self.colors['light'],
                padx=4,
                pady=2,
                cursor='hand2'
            ).pack(side=tk.LEFT, padx=5)
            
            if self.cloudtalk_api:
                tk.Button(
                    phone_frame,
                    text="📞 Call",
                    command=lambda p=phone, n=name: self._call_from_ai(p, n),
                    font=('Arial', 7),
                    bg=self.colors['success'],
                    fg='white',
                    padx=4,
                    pady=2,
                    cursor='hand2'
                ).pack(side=tk.LEFT, padx=2)
        
        # Address
        address = contact.get('current_address', '')
        if address:
            tk.Label(
                inner,
                text=f"📍 {address}",
                font=('Arial', 8),
                bg='#f8f9fa',
                fg='#666'
            ).pack(anchor='w', pady=2)
        
        # Source
        source = contact.get('source', '')
        if source:
            tk.Label(
                inner,
                text=f"Source: {source}",
                font=('Arial', 8, 'italic'),
                bg='#f8f9fa',
                fg='#999'
            ).pack(anchor='w', pady=2)
        
        # Reasoning
        reasoning = contact.get('reasoning', '')
        if reasoning:
            tk.Label(
                inner,
                text=f"Reasoning: {reasoning}",
                font=('Arial', 8),
                bg='#f8f9fa',
                fg='#666',
                wraplength=550,
                justify='left'
            ).pack(anchor='w', pady=2)
    
    def _copy_to_clipboard(self, text):
        """Copy text to clipboard"""
        self.root.clipboard_clear()
        self.root.clipboard_append(text)
        self.update_status(f"Copied: {text}", self.colors['success'])
    
    def _call_from_ai(self, phone, name):
        """Make call from AI tab"""
        if not self.cloudtalk_api or not self.agent_id:
            messagebox.showwarning("CloudTalk Not Configured", 
                                 "CloudTalk is not configured. Configure it in setup.")
            return
        
        # Track for call history
        self.last_called_phone = phone
        self.last_called_name = name
        
        # Make call in background
        def call_worker():
            result = self.cloudtalk_api.make_call(self.agent_id, phone)
            self.root.after(0, lambda: self._handle_call_result(result, phone, name))
        
        thread = threading.Thread(target=call_worker, daemon=True)
        thread.start()
        
        self.update_status(f"Calling {phone}...", self.colors['secondary'])
    
    def _handle_call_result(self, result, phone, name):
        """Handle call result from AI tab"""
        if result['success']:
            self.update_status(f"Call initiated to {phone}", self.colors['success'])
        else:
            self.update_status(f"Call failed: {result['message']}", self.colors['danger'])
            messagebox.showerror("Call Failed", result['message'])


def main():
    root = tk.Tk()
    app = DialerGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()
