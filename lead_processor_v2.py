#!/usr/bin/env python3
"""
TrestleIQ Lead Processor V2 - Two-Mode System
MODE 1: Bulk Processing Mode
MODE 2: Professional Dialer GUI Mode
"""

import pandas as pd
import requests
import time
import re
import base64
import os
from typing import Dict, List, Optional, Tuple
import logging
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.datavalidation import DataValidation
from cache_manager import CacheManager

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class LeadProcessor:
    def __init__(self, api_key: str, use_cache: bool = True, ai_assistant=None):
        self.api_key = api_key
        self.base_url = "https://api.trestleiq.com"
        self.headers = {
            'accept': 'application/json',
            'x-api-key': api_key
        }
        self.session = requests.Session()
        self.session.headers.update(self.headers)
        
        # Initialize cache manager
        self.use_cache = use_cache
        self.cache_manager = CacheManager() if use_cache else None
        
        # AI assistant for address correction
        self.ai_assistant = ai_assistant
        self.ai_correction_attempts = {}
        
    def parse_address(self, address_str: str) -> Dict[str, str]:
        """
        Parse a full address string into components for the API
        Handles various formats:
        - "1624 Clementson Drive San Antonio TX 78258 USA"
        - "123 Main St, Austin, TX 78701"
        - "456 Oak Avenue Apt 2B Houston TX 77001"
        - "789 Elm Street, Suite 100, Dallas, Texas 75201"
        Returns: {'street': '...', 'city': '...', 'state': '...', 'zip': '...'}
        """
        if not address_str or pd.isna(address_str):
            return {'street': '', 'city': '', 'state': '', 'zip': ''}
        
        address_str = str(address_str).strip()
        original_address = address_str
        
        # Remove country code if present (USA, US, United States, etc.)
        address_str = re.sub(r'\s*,?\s*(USA|US|United States|U\.S\.A\.|U\.S\.)$', '', address_str, flags=re.IGNORECASE)
        
        # Normalize commas - remove extra spaces around them
        address_str = re.sub(r'\s*,\s*', ', ', address_str)
        
        # Extract ZIP code (5 digits or 5+4 format)
        zip_pattern = r'\b(\d{5}(?:-\d{4})?)\b'
        zip_match = re.search(zip_pattern, address_str)
        
        if not zip_match:
            # No ZIP found, try to extract what we can
            logger.warning(f"No ZIP code found in address: {original_address}")
            return {'street': address_str, 'city': '', 'state': '', 'zip': ''}
        
        zip_code = zip_match.group(1)
        before_zip = address_str[:zip_match.start()].strip().rstrip(',')
        
        # State code or full state name before ZIP
        # Try 2-letter state code first
        state_pattern = r'\b([A-Z]{2})\s*$'
        state_match = re.search(state_pattern, before_zip)
        
        state_code = None
        if state_match:
            state_code = state_match.group(1)
            before_state = before_zip[:state_match.start()].strip().rstrip(',')
        else:
            # Try full state names
            state_names = {
                'alabama': 'AL', 'alaska': 'AK', 'arizona': 'AZ', 'arkansas': 'AR', 'california': 'CA',
                'colorado': 'CO', 'connecticut': 'CT', 'delaware': 'DE', 'florida': 'FL', 'georgia': 'GA',
                'hawaii': 'HI', 'idaho': 'ID', 'illinois': 'IL', 'indiana': 'IN', 'iowa': 'IA',
                'kansas': 'KS', 'kentucky': 'KY', 'louisiana': 'LA', 'maine': 'ME', 'maryland': 'MD',
                'massachusetts': 'MA', 'michigan': 'MI', 'minnesota': 'MN', 'mississippi': 'MS',
                'missouri': 'MO', 'montana': 'MT', 'nebraska': 'NE', 'nevada': 'NV', 'new hampshire': 'NH',
                'new jersey': 'NJ', 'new mexico': 'NM', 'new york': 'NY', 'north carolina': 'NC',
                'north dakota': 'ND', 'ohio': 'OH', 'oklahoma': 'OK', 'oregon': 'OR', 'pennsylvania': 'PA',
                'rhode island': 'RI', 'south carolina': 'SC', 'south dakota': 'SD', 'tennessee': 'TN',
                'texas': 'TX', 'utah': 'UT', 'vermont': 'VT', 'virginia': 'VA', 'washington': 'WA',
                'west virginia': 'WV', 'wisconsin': 'WI', 'wyoming': 'WY'
            }
            
            for state_name, code in state_names.items():
                pattern = r'\b' + re.escape(state_name) + r'\b\s*$'
                match = re.search(pattern, before_zip, flags=re.IGNORECASE)
                if match:
                    state_code = code
                    before_state = before_zip[:match.start()].strip().rstrip(',')
                    break
            
            if not state_code:
                # No state found
                logger.warning(f"No state found in address: {original_address}")
                return {'street': before_zip, 'city': '', 'state': '', 'zip': zip_code}
        
        # Now split remaining into street and city
        # Remove any leading/trailing commas
        before_state = before_state.strip().rstrip(',').lstrip(',').strip()
        
        # Split by comma if present
        if ',' in before_state:
            parts = [p.strip() for p in before_state.split(',')]
            # Last part before state is likely city
            if len(parts) >= 2:
                city = parts[-1]
                street = ', '.join(parts[:-1])
            else:
                # Only one part, assume it's street
                street = parts[0]
                city = ''
        else:
            # No comma, need to parse by words
            parts = before_state.split()
            
            if len(parts) <= 2:
                # Too short, assume it's all street
                return {'street': before_state, 'city': '', 'state': state_code, 'zip': zip_code}
            
            # Look for common street suffixes and secondary designators
            street_suffixes = [
                'street', 'st', 'avenue', 'ave', 'road', 'rd', 'drive', 'dr', 
                'lane', 'ln', 'court', 'ct', 'circle', 'cir', 'boulevard', 'blvd',
                'way', 'place', 'pl', 'parkway', 'pkwy', 'trail', 'terrace', 'ter',
                'highway', 'hwy', 'freeway', 'expressway', 'loop', 'path', 'pike',
                'row', 'run', 'square', 'sq', 'alley', 'walk', 'crossing'
            ]
            
            secondary_designators = [
                'apartment', 'apt', 'suite', 'ste', 'unit', 'building', 'bldg',
                'floor', 'fl', 'room', 'rm', '#', 'number', 'no'
            ]
            
            street_end_idx = None
            
            # First, look for secondary designators (Apt, Suite, etc.)
            for i, part in enumerate(parts):
                if part.lower().rstrip('.#') in secondary_designators:
                    # Everything from here to the end is part of street
                    # City comes after this
                    # Actually, secondary designators are part of street, so continue
                    pass
            
            # Look for street suffix
            for i, part in enumerate(parts):
                if part.lower().rstrip('.') in street_suffixes:
                    # Check if next word is a secondary designator
                    if i + 1 < len(parts) and parts[i + 1].lower().rstrip('.#') in secondary_designators:
                        # Include secondary designator and its value
                        if i + 2 < len(parts):
                            street_end_idx = i + 3  # Include suffix, designator, and value
                        else:
                            street_end_idx = i + 2
                    else:
                        street_end_idx = i + 1
                    break
            
            if street_end_idx and street_end_idx < len(parts):
                street = ' '.join(parts[:street_end_idx])
                city = ' '.join(parts[street_end_idx:])
            elif street_end_idx:
                # Street suffix was at the end
                street = ' '.join(parts)
                city = ''
            else:
                # No suffix found, use heuristic
                # Check if first word is a number (typical for street addresses)
                if parts[0].replace('#', '').replace('-', '').isdigit():
                    # Likely format: [Number] [Street Name...] [City Name...]
                    # City is usually 1-3 words at the end
                    if len(parts) >= 5:
                        # Assume last 2 words are city
                        street = ' '.join(parts[:-2])
                        city = ' '.join(parts[-2:])
                    elif len(parts) >= 4:
                        # Assume last 1-2 words are city
                        street = ' '.join(parts[:-2])
                        city = ' '.join(parts[-2:])
                    else:
                        # Assume last word is city
                        street = ' '.join(parts[:-1])
                        city = parts[-1]
                else:
                    # Doesn't start with number, harder to parse
                    # Default: last 1-2 words are city
                    if len(parts) >= 4:
                        street = ' '.join(parts[:-2])
                        city = ' '.join(parts[-2:])
                    else:
                        street = ' '.join(parts[:-1])
                        city = parts[-1]
        
        result = {
            'street': street.strip(),
            'city': city.strip(),
            'state': state_code,
            'zip': zip_code
        }
        
        # Log the parsing result
        logger.info(f"Parsed address: '{original_address}' -> Street: '{result['street']}', City: '{result['city']}', State: '{result['state']}', ZIP: '{result['zip']}'")
        
        return result
    
    def clean_phone(self, phone) -> str:
        """Clean and format phone number"""
        if pd.isna(phone):
            return ""
            
        phone_str = str(phone).strip()
        phone_clean = re.sub(r'[^\d+]', '', phone_str)
        
        if not phone_clean.startswith('+'):
            if phone_clean.startswith('1') and len(phone_clean) == 11:
                phone_clean = '+' + phone_clean
            elif len(phone_clean) == 10:
                phone_clean = '+1' + phone_clean
            elif len(phone_clean) == 11 and not phone_clean.startswith('1'):
                phone_clean = '+' + phone_clean
        
        return phone_clean
    
    def phone_lookup(self, phone: str) -> Dict:
        """Perform reverse phone lookup"""
        if not phone:
            return {}
            
        url = f"{self.base_url}/3.2/phone"
        params = {'phone': phone}
        
        try:
            response = self.session.get(url, params=params)
            response.raise_for_status()
            return response.json()
        except requests.exceptions.RequestException as e:
            logger.error(f"Phone lookup failed for {phone}: {e}")
            return {}
    
    def address_lookup(self, street: str, city: str, state: str, zip_code: str, 
                      enable_ai_correction: bool = True, status_callback=None) -> Tuple[Dict, Optional[Dict]]:
        """
        Perform reverse address lookup with optional AI correction
        Returns: (api_response, correction_info)
        correction_info contains: {'original': {...}, 'corrected': {...}, 'reasoning': '...'}
        """
        if not any([street, city, state, zip_code]):
            return {}, None
            
        url = f"{self.base_url}/3.1/location"
        params = {}
        
        if street:
            params['street_line_1'] = street
        if city:
            params['city'] = city
        if state:
            params['state_code'] = state
        if zip_code:
            params['postal_code'] = str(zip_code).split('-')[0]
        
        # Create cache key for AI correction tracking
        cache_key = f"{street}_{city}_{state}_{zip_code}"
        
        # First attempt
        try:
            if status_callback:
                status_callback("Looking up address...")
            
            response = self.session.get(url, params=params)
            response.raise_for_status()
            result = response.json()
            
            # Check if lookup was successful
            is_valid = result.get('is_valid', False)
            has_residents = bool(result.get('current_residents', []))
            
            if is_valid and has_residents:
                return result, None
            
            # If failed and AI is enabled, try correction
            if enable_ai_correction and self.ai_assistant:
                # Check if we've already tried correcting this address
                if cache_key in self.ai_correction_attempts:
                    attempts = self.ai_correction_attempts[cache_key]
                    if attempts >= 2:
                        logger.info(f"Max AI correction attempts reached for {cache_key}")
                        return result, None
                else:
                    self.ai_correction_attempts[cache_key] = 0
                
                # Try AI correction
                if status_callback:
                    status_callback("Address lookup failed - AI correcting format...")
                
                logger.info("Attempting AI address correction...")
                corrected = self.ai_assistant.correct_address(
                    street_line_1=street or "",
                    street_line_2="",
                    city=city or "",
                    state_code=state or "",
                    postal_code=zip_code or ""
                )
                
                if corrected:
                    self.ai_correction_attempts[cache_key] += 1
                    
                    # Retry with corrected address
                    if status_callback:
                        status_callback(f"Retrying with corrected address (attempt {self.ai_correction_attempts[cache_key]})...")
                    
                    corrected_params = {
                        'street_line_1': corrected.get('street_line_1', ''),
                        'city': corrected.get('city', ''),
                        'state_code': corrected.get('state_code', ''),
                        'postal_code': corrected.get('postal_code', '')
                    }
                    
                    # Remove empty params
                    corrected_params = {k: v for k, v in corrected_params.items() if v}
                    
                    logger.info(f"AI correction: {params} -> {corrected_params}")
                    
                    retry_response = self.session.get(url, params=corrected_params)
                    retry_response.raise_for_status()
                    retry_result = retry_response.json()
                    
                    if retry_result.get('is_valid') and retry_result.get('current_residents'):
                        if status_callback:
                            status_callback("Address corrected and verified ✓")
                        
                        correction_info = {
                            'original': params,
                            'corrected': corrected_params,
                            'reasoning': corrected.get('correction_reasoning', 'AI correction applied')
                        }
                        return retry_result, correction_info
                    else:
                        # If still failed and we haven't hit max attempts, try one more time
                        if self.ai_correction_attempts[cache_key] < 2:
                            if status_callback:
                                status_callback("Address correction attempt 2...")
                            
                            # Recursive call for second attempt
                            return self.address_lookup(street, city, state, zip_code, 
                                                     enable_ai_correction=True, 
                                                     status_callback=status_callback)
            
            return result, None
            
        except requests.exceptions.RequestException as e:
            logger.error(f"Address lookup failed: {e}")
            return {}, None
    
    def extract_age(self, person_data: Dict) -> str:
        """Extract age or age_range from person data"""
        if 'age' in person_data and person_data['age']:
            return str(person_data['age'])
        if 'age_range' in person_data and person_data['age_range']:
            return person_data['age_range']
        return ""
    
    def get_last_name(self, full_name: str) -> str:
        """Extract last name from full name"""
        if not full_name or pd.isna(full_name):
            return ""
        parts = str(full_name).strip().split()
        if len(parts) > 0:
            return parts[-1].upper()
        return ""
    
    def format_address(self, address_data: Dict) -> str:
        """Format address from address data"""
        if not address_data:
            return ""
        parts = []
        if address_data.get('street_line_1'):
            parts.append(address_data['street_line_1'])
        if address_data.get('city'):
            parts.append(address_data['city'])
        if address_data.get('state_code'):
            parts.append(address_data['state_code'])
        if address_data.get('postal_code'):
            parts.append(address_data['postal_code'])
        return ', '.join(parts)
    
    def process_row(self, row: pd.Series, cached_data: Optional[Dict] = None, force_refresh: bool = False) -> Tuple[List[Dict], bool]:
        """
        Process a single row and return list of result rows and address_error flag
        Returns: (results, address_lookup_failed)
        If cached_data provided, skip API calls and use cached data
        If force_refresh is True, bypass all caches and make fresh API calls
        """
        logger.info(f"Processing: {row.get('name', 'Unknown')}")
        
        # Check activity score
        activity_score = row.get('activity_score', row.get('activity score', 0))
        if pd.notna(activity_score) and float(activity_score) > 30:
            logger.info(f"Skipping - activity score {activity_score} > 30")
            return [], False
        
        original_address = row.get('address', '')
        original_phone = self.clean_phone(row.get('phone', ''))
        original_name = row.get('name', '')
        
        results = []
        address_lookup_failed = False
        
        # Use cached data if provided (from Excel cache) - unless force refresh
        if cached_data and not force_refresh:
            logger.info("Using cached data from output Excel")
            return [cached_data], cached_data.get('address_lookup_failed', False)
        
        # Check permanent cache first - unless force refresh
        phone_data = None
        address_data = None
        used_cache = False
        
        if not force_refresh and self.use_cache and self.cache_manager:
            cached_lookup = self.cache_manager.get_cached_lookup(original_phone)
            if cached_lookup:
                phone_data, address_data, _ = cached_lookup  # Ignore AI analysis in lead processor
                used_cache = True
                logger.info(f"Using permanent cache for {original_phone}")
        
        # If not in cache or force refresh, make API calls
        if not used_cache or force_refresh:
            # Phone lookup
            phone_data = self.phone_lookup(original_phone)
            time.sleep(0.1)
            
            # Address lookup - parse address if needed
            # Check if we have separate address fields or a single address string
            street = row.get('address', '')
            city = row.get('city', '')
            state = row.get('state', '')
            zip_code = row.get('zip', '')
            
            # If city/state/zip are missing but we have an address, try to parse it
            if street and (not city or not state or not zip_code):
                logger.info(f"Parsing combined address: {street}")
                parsed = self.parse_address(street)
                street = parsed['street'] or street
                city = parsed['city'] or city
                state = parsed['state'] or state
                zip_code = parsed['zip'] or zip_code
                logger.info(f"Parsed -> Street: {street}, City: {city}, State: {state}, ZIP: {zip_code}")
            
            address_data = self.address_lookup(street, city, state, zip_code)
            time.sleep(0.1)
            
            # Store in permanent cache if we have data (always store on force refresh to update cache)
            if self.use_cache and self.cache_manager and (phone_data or address_data):
                self.cache_manager.store_lookup(original_phone, phone_data or {}, address_data or {})
                if force_refresh:
                    logger.info(f"Updated cache with fresh API data for {original_phone}")
                else:
                    logger.info(f"Stored lookup in permanent cache for {original_phone}")
        
        # Check if address lookup returned no results or error
        if not address_data or 'current_residents' not in address_data or not address_data['current_residents']:
            address_lookup_failed = True
        
        # Collect all found data
        found_people = []
        
        # From phone lookup
        if phone_data and 'owners' in phone_data:
            owners = phone_data.get('owners', [])
            if owners is None:
                owners = []
            
            for owner in owners:
                if not isinstance(owner, dict):
                    continue
                
                person_info = {
                    'name': owner.get('name', ''),
                    'age': self.extract_age(owner),
                    'phones': [],
                    'addresses': []
                }
                
                if 'phone_number' in phone_data:
                    person_info['phones'].append(phone_data['phone_number'])
                
                for alt_phone in owner.get('alternate_phones', []) or []:
                    phone_num = None
                    if isinstance(alt_phone, dict):
                        # Try different possible keys
                        phone_num = alt_phone.get('phone_number') or alt_phone.get('phoneNumber') or alt_phone.get('phone')
                    elif isinstance(alt_phone, str):
                        # Sometimes it's just a string
                        phone_num = alt_phone
                    
                    if phone_num and phone_num not in person_info['phones']:
                        person_info['phones'].append(phone_num)
                
                for addr in owner.get('current_addresses', []) or []:
                    if isinstance(addr, dict):
                        formatted_addr = self.format_address(addr)
                        if formatted_addr and formatted_addr not in person_info['addresses']:
                            person_info['addresses'].append(formatted_addr)
                
                found_people.append(person_info)
        
        # From address lookup
        if address_data and 'current_residents' in address_data:
            residents = address_data.get('current_residents', [])
            if residents is None:
                residents = []
            
            for resident in residents:
                if not isinstance(resident, dict):
                    continue
                
                resident_name = resident.get('name', '')
                existing = next((p for p in found_people if p['name'] == resident_name), None)
                
                if existing:
                    person_info = existing
                else:
                    person_info = {
                        'name': resident_name,
                        'age': self.extract_age(resident),
                        'phones': [],
                        'addresses': []
                    }
                    found_people.append(person_info)
                
                for phone in resident.get('phones', []) or []:
                    phone_num = None
                    if isinstance(phone, dict):
                        # Try different possible keys
                        phone_num = phone.get('phone_number') or phone.get('phoneNumber') or phone.get('phone')
                    elif isinstance(phone, str):
                        # Sometimes it's just a string
                        phone_num = phone
                    
                    if phone_num and phone_num not in person_info['phones']:
                        person_info['phones'].append(phone_num)
                
                formatted_addr = self.format_address(address_data)
                if formatted_addr and formatted_addr not in person_info['addresses']:
                    person_info['addresses'].append(formatted_addr)
        
        # Create output rows - ONE ROW PER PERSON with all phones and addresses
        if not found_people:
            results.append({
                'original_address': original_address,
                'original_phone': original_phone,
                'blank': '',
                'age': '',
                'original_name': original_name,
                'new_phones': [],  # List of all phones
                'new_addresses': [],  # List of all addresses
                'new_name': 'NO RESULTS FOUND',
                'status': '',
                'address_lookup_failed': address_lookup_failed
            })
        else:
            for person in found_people:
                results.append({
                    'original_address': original_address,
                    'original_phone': original_phone,
                    'blank': '',
                    'age': person['age'],
                    'original_name': original_name,
                    'new_phones': person['phones'],  # All phones as list
                    'new_addresses': person['addresses'],  # All addresses as list
                    'new_name': person['name'],
                    'status': '',
                    'address_lookup_failed': address_lookup_failed
                })
        
        return results, address_lookup_failed
    
    def process_excel_file_bulk(self, input_file: str, output_file: str, max_rows: Optional[int] = None, 
                                progress_callback=None, status_callback=None):
        """MODE 1: Bulk Processing Mode with progress tracking and caching"""
        logger.info(f"Reading Excel file: {input_file}")
        
        df = pd.read_excel(input_file)
        total_rows = len(df)
        
        if max_rows:
            total_rows = min(max_rows, total_rows)
            df = df.head(max_rows)
        
        logger.info(f"Processing {total_rows} rows")
        
        # Load cache from existing output file
        cache = self._load_cache_from_excel(output_file)
        
        all_results = []
        rows_with_address_errors = []
        cached_count = 0
        api_count = 0
        
        for index, row in df.iterrows():
            if progress_callback:
                progress_callback(index + 1, total_rows)
            
            original_phone = self.clean_phone(row.get('phone', ''))
            
            # Check cache first
            if original_phone in cache:
                if status_callback:
                    status_callback(f"Row {index + 1}/{total_rows} - Loading from cache")
                cached_data = cache[original_phone]
                results = [cached_data]
                address_error = cached_data.get('address_lookup_failed', False)
                cached_count += 1
            else:
                if status_callback:
                    status_callback(f"Row {index + 1}/{total_rows} - Making API calls")
                results, address_error = self.process_row(row)
                api_count += 1
            
            if address_error:
                start_idx = len(all_results)
                all_results.extend(results)
                end_idx = len(all_results)
                rows_with_address_errors.extend(range(start_idx, end_idx))
            else:
                all_results.extend(results)
        
        # Create Excel file with formatting
        self.create_excel_output(all_results, output_file, rows_with_address_errors)
        
        logger.info(f"Processing complete! Results saved to: {output_file}")
        logger.info(f"Excel cache hits: {cached_count}, API calls: {api_count}")
        
        # Log permanent cache statistics
        if self.use_cache and self.cache_manager:
            stats = self.cache_manager.get_statistics()
            logger.info(f"Permanent cache - Hits: {stats['cache_hits']}, Misses: {stats['cache_misses']}, Hit rate: {stats['hit_rate']}%")
            logger.info(f"API calls saved: {stats['api_calls_saved']}, Estimated savings: ${stats['estimated_savings_usd']}")
        
        return len(all_results), cached_count, api_count
    
    def _load_cache_from_excel(self, output_file: str) -> Dict:
        """Load cache from existing output Excel file"""
        cache = {}
        if not os.path.exists(output_file):
            return cache
        
        try:
            df = pd.read_excel(output_file)
            for _, row in df.iterrows():
                orig_phone = str(row.get('Original Phone', '')).strip()
                if orig_phone:
                    # Split phones by comma, addresses by pipe
                    phones_raw = str(row.get('New Phone', ''))
                    addresses_raw = str(row.get('New Address', ''))
                    
                    cache[orig_phone] = {
                        'original_address': row.get('Original Address', ''),
                        'original_phone': orig_phone,
                        'blank': '',
                        'age': row.get('Age', ''),
                        'original_name': row.get('Original Name', ''),
                        'new_phones': [p.strip() for p in phones_raw.split(',') if p.strip()] if pd.notna(row.get('New Phone')) else [],
                        'new_addresses': [a.strip() for a in addresses_raw.split('|') if a.strip()] if pd.notna(row.get('New Address')) else [],
                        'new_name': row.get('New Name', ''),
                        'status': row.get('Status', ''),
                        'address_lookup_failed': False
                    }
        except Exception as e:
            logger.error(f"Error loading cache: {e}")
        
        return cache
    
    def create_excel_output(self, results: List[Dict], output_file: str, red_row_indices: List[int] = None):
        """Create Excel file with specific formatting - ONE ROW PER PERSON"""
        if red_row_indices is None:
            red_row_indices = []
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Lead Results"
        
        headers = [
            'Original Address',
            'Original Phone',
            '',
            'Age',
            'Original Name',
            'New Phone',
            'New Address',
            'New Name',
            'Status',
            'Notes'
        ]
        
        ws.append(headers)
        
        # Style header row
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Define fills
        green_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        
        # Write data rows
        for idx, result in enumerate(results):
            # Format phones and addresses for display
            phones_list = result.get('new_phones', [])
            addresses_list = result.get('new_addresses', [])
            
            # Join all phones with comma, all addresses with pipe
            phones_str = ", ".join(phones_list) if phones_list else ''
            addresses_str = " | ".join(addresses_list) if addresses_list else ''
            
            row_data = [
                result['original_address'],
                result['original_phone'],
                result['blank'],
                result['age'],
                result['original_name'],
                phones_str,
                addresses_str,
                result['new_name'],
                result['status'],
                result.get('notes', '')
            ]
            
            ws.append(row_data)
            current_row = ws.max_row
            
            # Check if this row should be red (address lookup failed)
            if result.get('address_lookup_failed', False):
                for cell in ws[current_row]:
                    cell.fill = red_fill
                    cell.font = Font(color="FFFFFF")
            else:
                # Check for last name match and highlight green
                original_last = self.get_last_name(result['original_name'])
                new_last = self.get_last_name(result['new_name'])
                
                if original_last and new_last and original_last == new_last:
                    for cell in ws[current_row]:
                        cell.fill = green_fill
            
            # Add phone dropdown if multiple phones
            if len(phones_list) > 1:
                phone_cell = ws.cell(row=current_row, column=6)
                # Use comma separator for phones in dropdown
                phone_dv = DataValidation(type="list", formula1=f'"{",".join(phones_list)}"', allow_blank=True)
                phone_dv.add(phone_cell)
                ws.add_data_validation(phone_dv)
            
            # No dropdown for addresses - just text separated by |
        
        # Add Status dropdown
        status_options = "No answer,Not interested,Wrong Number,Call back,Appointment,DNC"
        status_dv = DataValidation(type="list", formula1=f'"{status_options}"', allow_blank=True)
        
        if len(results) > 0:
            status_dv.add(f'I2:I{ws.max_row}')  # Column I is Status (column 9)
            ws.add_data_validation(status_dv)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        ws.column_dimensions['C'].width = 2
        
        wb.save(output_file)


def main_bulk_mode():
    """MODE 1: Bulk Processing Mode"""
    import sys
    
    try:
        from config import API_KEY
    except ImportError:
        print("Error: config.py not found or API_KEY not defined")
        sys.exit(1)
    
    if not API_KEY:
        print("Error: API_KEY is empty in config.py")
        sys.exit(1)
    
    # Get input file
    if len(sys.argv) > 1:
        input_file = sys.argv[1]
    else:
        input_file = input("Enter input Excel file path: ").strip()
    
    if not input_file:
        print("Error: No input file specified")
        sys.exit(1)
    
    # Ask for number of rows to process
    max_rows_input = input("Enter number of rows to process (press Enter for all): ").strip()
    max_rows = None
    if max_rows_input:
        try:
            max_rows = int(max_rows_input)
        except ValueError:
            print("Invalid number, processing all rows")
    
    # Get output file
    if len(sys.argv) > 2:
        output_file = sys.argv[2]
    else:
        output_file = input("Enter output Excel file path (default: leads_processed.xlsx): ").strip()
        if not output_file:
            output_file = "leads_processed.xlsx"
    
    if not output_file.endswith('.xlsx'):
        output_file += '.xlsx'
    
    print("\nTrestleIQ Lead Processor - BULK MODE")
    print("=" * 50)
    print(f"Input file: {input_file}")
    print(f"Output file: {output_file}")
    if max_rows:
        print(f"Processing: {max_rows} rows")
    else:
        print("Processing: All rows")
    print()
    
    processor = LeadProcessor(API_KEY)
    
    try:
        num_results = processor.process_excel_file_bulk(input_file, output_file, max_rows)
        print(f"\n✓ Processing complete!")
        print(f"✓ Total results: {num_results}")
        print(f"✓ Output saved to: {output_file}")
        print("\nFormatting:")
        print("  • RED rows: Address lookup returned no results or errors")
        print("  • GREEN rows: Last name matches between original and new name")
        print("  • Status dropdown available in last column")
    except Exception as e:
        print(f"\n✗ Error: {e}")
        logger.exception("Processing failed")
        sys.exit(1)


if __name__ == "__main__":
    main_bulk_mode()
